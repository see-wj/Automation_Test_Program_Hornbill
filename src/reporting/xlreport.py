import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
import pandas as pd
import datetime
import os
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
import win32com.client 
import re
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl import load_workbook
from openpyxl.chart.legend import Legend
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
)
from common.path import csv_folder, IMAGE_DIR, IMAGE_PATH, IMAGE_PATH_2


def configure_run_storage(raw_directory, chart_directory):
    global csv_folder, IMAGE_DIR, IMAGE_PATH, IMAGE_PATH_2
    csv_folder = raw_directory
    IMAGE_DIR = chart_directory
    IMAGE_PATH = chart_directory / "Chart.png"
    IMAGE_PATH_2 = chart_directory / "Chart2.png"


class xlreport:
    """Generates an Excel report for voltage/current accuracy tests with conditional formatting and embedded images."""

    CHART_SERIES_COLORS = (
        "4472C4",
        "ED7D31",
        "7030A0",
        "00A6A6",
        "A0522D",
        "C55A11",
        "5B9BD5",
        "8064A2",
    )

    def __init__(
        self,
        save_directory,
        file_name=None,
        raw_directory=None,
        chart_directory=None,
    ):
        """Initialize formatting parameters and set file path for Excel report."""
        self.red_font = Font(size=14, bold=True, color="ffffff")
        self.red_fill = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")
        self.green_font = Font(size=14, bold=True, color="013220")
        self.green_fill = PatternFill(start_color="FFAAFF00", end_color="FFAAFF00", fill_type="solid")
        run_root = Path(save_directory).parent
        inferred_raw = run_root / "raw"
        inferred_charts = run_root / "charts"
        self.csv_folder = Path(raw_directory) if raw_directory else (
            inferred_raw if inferred_raw.exists() else Path(csv_folder)
        )
        self.chart_directory = Path(chart_directory) if chart_directory else (
            inferred_charts if inferred_charts.exists() else Path(IMAGE_DIR)
        )

        # Ensure the save directory exists
        if not os.path.exists(save_directory):
            os.makedirs(save_directory)

        # Generate a timestamped filename
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d--%H-%M-%S")
        if file_name:
            full_file_name = f"{file_name}_{timestamp}.xlsx"
        else:
            full_file_name = f"{timestamp}.xlsx"

        self.path = os.path.join(save_directory, full_file_name)

    def adjust_column_width(self, worksheet, width=20):
        """Set column width for columns A to R."""
        for col in range(1, 25):  # A to X
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    def align_cell(self, worksheet, row, col, style="center"):
        """Align a specific cell in the worksheet."""
        worksheet.cell(row=row, column=col).alignment = Alignment(horizontal=style)

    def _add_condition_formatting(
        self,
        worksheet,
        dataframe,
        start_column=4,
        first_data_row=9,
    ):
        last_data_row = first_data_row + len(dataframe) - 1
        for column_name in (
            "Programming Condition",
            "Readback Condition",
            "Programming Percentage Condition",
            "Readback Percentage Condition",
        ):
            if column_name not in dataframe.columns:
                continue
            column_number = start_column + dataframe.columns.get_loc(column_name)
            column_letter = get_column_letter(column_number)
            cell_range = (
                f"{column_letter}{first_data_row}:"
                f"{column_letter}{last_data_row}"
            )
            first_cell = f"{column_letter}{first_data_row}"
            worksheet.conditional_formatting.add(
                cell_range,
                FormulaRule(
                    formula=[f'NOT(ISERROR(SEARCH("PASS",{first_cell})))'],
                    fill=self.green_fill,
                    font=self.green_font,
                ),
            )
            worksheet.conditional_formatting.add(
                cell_range,
                FormulaRule(
                    formula=[f'NOT(ISERROR(SEARCH("FAIL",{first_cell})))'],
                    fill=self.red_fill,
                    font=self.red_font,
                ),
            )

    @staticmethod
    def _bold_text_properties(size):
        character_properties = CharacterProperties(
            b=True,
            sz=size,
            solidFill="000000",
        )
        return RichText(
            p=[
                Paragraph(
                    pPr=ParagraphProperties(
                        defRPr=character_properties,
                    ),
                    endParaRPr=CharacterProperties(
                        b=True,
                        sz=size,
                        solidFill="000000",
                    ),
                )
            ]
        )

    @staticmethod
    def _bold_chart_title(title, size):
        if title is None or title.tx is None or title.tx.rich is None:
            return
        for paragraph in title.tx.rich.p:
            paragraph.pPr = ParagraphProperties(
                defRPr=CharacterProperties(
                    b=True,
                    sz=size,
                    solidFill="000000",
                )
            )
            paragraph.endParaRPr = CharacterProperties(
                b=True,
                sz=size,
                solidFill="000000",
            )
            for run in paragraph.r:
                run.rPr = CharacterProperties(
                    b=True,
                    sz=size,
                    solidFill="000000",
                )

    def _style_native_chart(self, chart):
        self._bold_chart_title(chart.title, 1600)
        self._bold_chart_title(chart.x_axis.title, 1200)
        self._bold_chart_title(chart.y_axis.title, 1200)
        chart.x_axis.txPr = self._bold_text_properties(1000)
        chart.y_axis.txPr = self._bold_text_properties(1000)
        chart.legend.txPr = self._bold_text_properties(1000)

    @staticmethod
    def _chart_specs(dataframe):
        mode = str(dataframe["Chart Mode"].iloc[0])
        if mode.startswith("VOLTAGE"):
            x_title = (
                "Current (A)"
                if mode == "VOLTAGE_CURRENT_CHANGE"
                else "Voltage (V)"
            )
            chart_specs = [
                (
                    "Voltage Programming Error",
                    "Programming/Voltage Absolute Error (V)",
                    "Programming Upper Error Boundary (V)",
                    "Programming Lower Error Boundary (V)",
                    "Programming Condition",
                    "Programming/Voltage Absolute Error (V)",
                ),
                (
                    "Voltage Readback Error",
                    "PSU Readback Voltage Error (V)",
                    "Readback Upper Error Boundary (V)",
                    "Readback Lower Error Boundary (V)",
                    "Readback Condition",
                    "PSU Readback Voltage Error (V)",
                ),
                (
                    "Voltage Programming Percentage Error",
                    "Relative/Voltage Percentage Error (%)",
                    "Programming Upper Percentage Error Boundary (%)",
                    "Programming Lower Percentage Error Boundary (%)",
                    "Programming Percentage Condition",
                    "Relative/Voltage Percentage Error (%)",
                ),
                (
                    "Voltage Readback Percentage Error",
                    "PSU Readback Voltage Percentage Error (%)",
                    "Readback Upper Percentage Error Boundary (%)",
                    "Readback Lower Percentage Error Boundary (%)",
                    "Readback Percentage Condition",
                    "PSU Readback Voltage Percentage Error (%)",
                ),
            ]
            if "PSU Local Voltage Error (V)" in dataframe.columns:
                chart_specs.append(
                    (
                        "PSU Local Voltage Error (VLOC - VMON)",
                        "PSU Local Voltage Error (V)",
                        None,
                        None,
                        None,
                        "PSU Local Voltage Error (V)",
                    )
                )
            return x_title, tuple(chart_specs)

        return "Current (A)", (
            (
                "Current Programming Error",
                "Programming/Current Absolute Error (A)",
                "Programming Upper Error Boundary (A)",
                "Programming Lower Error Boundary (A)",
                "Programming Condition",
                "Programming/Current Absolute Error (A)",
            ),
            (
                "Current Readback Error",
                "PSU Readback Current Error (A)",
                "Readback Upper Error Boundary (A)",
                "Readback Lower Error Boundary (A)",
                "Readback Condition",
                "PSU Readback Current Error (A)",
            ),
            (
                "Current Programming Percentage Error",
                "Relative/Current Percentage Error (%)",
                "Programming Upper Percentage Error Boundary (%)",
                "Programming Lower Percentage Error Boundary (%)",
                "Programming Percentage Condition",
                "Relative/Current Percentage Error (%)",
            ),
            (
                "Current Readback Percentage Error",
                "PSU Readback Current Percentage Error (%)",
                "Readback Upper Percentage Error Boundary (%)",
                "Readback Lower Percentage Error Boundary (%)",
                "Readback Percentage Condition",
                "PSU Readback Current Percentage Error (%)",
            ),
        )

    @staticmethod
    def _write_chart_values(worksheet, column, title, x_values, y_values):
        worksheet.cell(row=1, column=column, value=f"{title} X")
        worksheet.cell(row=1, column=column + 1, value=title)
        for row, (x_value, y_value) in enumerate(zip(x_values, y_values), start=2):
            worksheet.cell(row=row, column=column, value=x_value)
            worksheet.cell(row=row, column=column + 1, value=y_value)
        return column, column + 1, len(x_values) + 1

    @staticmethod
    def _append_chart_series(
        chart,
        worksheet,
        columns,
        title,
        color,
        marker=False,
        line=True,
        dashed=False,
        line_width=None,
    ):
        x_column, y_column, last_row = columns
        if last_row < 2:
            return
        x_values = Reference(
            worksheet,
            min_col=x_column,
            min_row=2,
            max_row=last_row,
        )
        y_values = Reference(
            worksheet,
            min_col=y_column,
            min_row=2,
            max_row=last_row,
        )
        series = Series(y_values, x_values, title=title)
        if line:
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = pixels_to_EMU(
                line_width if line_width is not None else 3.2
            )
            if dashed:
                series.graphicalProperties.line.dashStyle = "dash"
        else:
            series.graphicalProperties.line.noFill = True
        if marker:
            series.marker.symbol = "circle"
            series.marker.size = 8
            series.marker.graphicalProperties.solidFill = color
            series.marker.graphicalProperties.line.solidFill = color
        else:
            series.marker.symbol = "none"
        chart.series.append(series)

    def _add_native_charts(self, workbook, dataframe):
        required = {"Chart Mode", "Chart X", "Chart Group"}
        if dataframe.empty or not required.issubset(dataframe.columns):
            return

        for sheet_name in ("Charts", "ChartData"):
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
        chart_sheet = workbook.create_sheet("Charts")
        source_sheet = workbook.create_sheet("ChartData")
        source_sheet.sheet_state = "hidden"
        x_title, chart_specs = self._chart_specs(dataframe)
        chart_positions = ("A1", "J1", "A22", "J22", "A43")
        source_column = 1

        for position, spec in zip(chart_positions, chart_specs):
            title, value_column, upper_column, lower_column, condition_column, y_title = spec
            required_columns = {value_column}
            required_columns.update(
                column
                for column in (upper_column, lower_column, condition_column)
                if column is not None
            )
            if not required_columns.issubset(dataframe.columns):
                continue

            chart = ScatterChart()
            chart.scatterStyle = "lineMarker"
            chart.style = 13
            chart.title = title
            chart.x_axis.title = x_title
            chart.y_axis.title = y_title
            chart.width = 18
            chart.height = 11.5
            chart.legend.position = "b"
            self._style_native_chart(chart)

            for group_index, (group_title, group) in enumerate(
                dataframe.groupby("Chart Group", sort=False)
            ):
                points = group[["Chart X", value_column]].dropna()
                columns = self._write_chart_values(
                    source_sheet,
                    source_column,
                    str(group_title),
                    points["Chart X"].tolist(),
                    points[value_column].tolist(),
                )
                self._append_chart_series(
                    chart,
                    source_sheet,
                    columns,
                    str(group_title),
                    self.CHART_SERIES_COLORS[
                        group_index % len(self.CHART_SERIES_COLORS)
                    ],
                )
                source_column += 2

            if upper_column is not None and lower_column is not None:
                boundary_points = dataframe[
                    ["Chart X", upper_column, lower_column]
                ].dropna()
                for boundary_title, boundary_column in (
                    ("Upper Bound", upper_column),
                    ("Lower Bound", lower_column),
                ):
                    columns = self._write_chart_values(
                        source_sheet,
                        source_column,
                        boundary_title,
                        boundary_points["Chart X"].tolist(),
                        boundary_points[boundary_column].tolist(),
                    )
                    self._append_chart_series(
                        chart,
                        source_sheet,
                        columns,
                        boundary_title,
                        "FF0000",
                        dashed=True,
                        line_width=4.5,
                    )
                    source_column += 2

            if condition_column is not None:
                for status, color in (("PASS", "00B050"), ("FAIL", "FF0000")):
                    status_points = dataframe.loc[
                        dataframe[condition_column].astype(str).str.upper() == status,
                        ["Chart X", value_column],
                    ].dropna()
                    columns = self._write_chart_values(
                        source_sheet,
                        source_column,
                        status,
                        status_points["Chart X"].tolist(),
                        status_points[value_column].tolist(),
                    )
                    self._append_chart_series(
                        chart,
                        source_sheet,
                        columns,
                        status,
                        color,
                        marker=True,
                        line=False,
                    )
                    source_column += 2

            x_values = pd.to_numeric(dataframe["Chart X"], errors="coerce").dropna()
            if not x_values.empty:
                columns = self._write_chart_values(
                    source_sheet,
                    source_column,
                    "Zero",
                    [float(x_values.min()), float(x_values.max())],
                    [0.0, 0.0],
                )
                self._append_chart_series(
                    chart,
                    source_sheet,
                    columns,
                    "Zero",
                    "808080",
                    dashed=True,
                )
                source_column += 2

            chart_sheet.add_chart(chart, position)

    def run(self):
        """Generate the Excel report by importing CSV data, adding conditional formatting, and embedding an image."""
        """# Define chart_path for image insertion
        current_dir = os.path.dirname(os.path.abspath(__file__))  # src/
        project_root = os.path.abspath(os.path.join(current_dir, ".."))
        csv_folder = os.path.join(project_root, "csv")
        chart_path = os.path.join(csv_folder, "images", "Chart.png")
        chart_path_2 = os.path.join(csv_folder, "images", "Chart2.png")"""
        csv_directory = self.csv_folder
        chart_path = str(self.chart_directory / "Chart.png")
        chart_path_2 = str(self.chart_directory / "Chart2.png")
        try:
            with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
                # Read CSV files
                try:
                    df1 = pd.read_csv(os.path.join(csv_directory, "error.csv"))
                    chart_dataframe = df1
                    percentage_path = os.path.join(
                        csv_directory,
                        "error_percent.csv",
                    )
                    if os.path.exists(percentage_path):
                        chart_dataframe = pd.read_csv(percentage_path)
                    #df3 = pd.read_csv(os.path.join(csv_folder, "error_percent.csv"))
                    df2 = pd.read_csv(os.path.join(csv_directory, "instrumentData.csv"))
                    df4 = pd.read_csv(os.path.join(csv_directory, "config.csv"))
                except FileNotFoundError as e:
                    print(f"Error: {e}")
                    return

                # Write DataFrames to Excel
                df1.to_excel(writer, sheet_name="Data", index=False, startrow=7, startcol=3)
                #df3.to_excel(writer, sheet_name="Data", index=False, startrow=7, startcol=16)
                df2.to_excel(writer, sheet_name="Data", index=False)
                df4.to_excel(writer, sheet_name="Data", index=False, startrow=6)

                # Access workbook and sheet
                wb = writer.book
                ws = wb["Data"]

                self._add_condition_formatting(ws, df1)
                self._add_native_charts(wb, chart_dataframe)


                # Adjust column widths
                self.adjust_column_width(ws, width=25)

                # Add timestamp to the sheet
                ws.cell(row=7, column=4, value="Time Generated:")
                ws.cell(row=7, column=5, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

                if os.path.exists(chart_path):
                    chart_image = openpyxl.drawing.image.Image(chart_path)
                    chart_image.anchor = "A34"
                    ws.add_image(chart_image)
                else:
                    print(
                        f"Warning: Image '{chart_path}' not found. "
                        "Skipping image insertion."
                    )

                if os.path.exists(chart_path_2):
                    percentage_chart_image = openpyxl.drawing.image.Image(chart_path_2)
                    percentage_chart_image.anchor = "A67"
                    ws.add_image(percentage_chart_image)
                else:
                    print(
                        f"Warning: Image '{chart_path_2}' not found. "
                        "Skipping image insertion."
                    )

                # Save the workbook
                wb.save(self.path)

            print(f"Excel report saved: {self.path}")

        except Exception as e:
            import traceback
            traceback.print_exc()



 
 
class Graph_Plotting:
    def __init__(self, workbook, sheet, output_file, keyword, num_test_running, \
                graph_start_rows, final_prog_df = None):
 
        self.wb = workbook
        self.ws = sheet
        self.output_file = output_file
        self.keyword = keyword
        self.num_test_running = num_test_running
        self.graph_start_rows = graph_start_rows
        self.final_prog_df = final_prog_df

       
        print("✅Plotting Graph...")
 
    def set_chart(self, graph_name, data_start_row, label_row, max_row, categories_programming, columns):
        chart1 = LineChart()
        chart1.width = 30
        chart1.height = 20
        chart1.style = 42
 
        # Add all data series first
        for name, col in columns.items():
            # Get label from first row
            label = self.ws.cell(row=label_row, column=col).value or name
            # Get data values starting from data row (skip empty headers)
            values = Reference(self.ws, min_col=col, min_row=data_start_row, max_row=max_row)
            series = Series(values, title=label)
            chart1.series.append(series)
   
        # Set axis and categories_programming
        chart1.set_categories(categories_programming)
        chart1.title = graph_name
        chart1.y_axis.title = "Percentage"
        chart1.x_axis.title = "Power"
        chart1.legend.position = "b"
 
        return chart1  
   
    def line_style(self, chart, colors):
        for series, color in zip(chart.series, colors):
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = pixels_to_EMU(2)
            if color == "FF0000":
                series.graphicalProperties.line.dashStyle = "dash"  # Dotted
            if color == "FFFFFF":
                series.graphicalProperties.line.dashStyle = "dash"  # Dotted
            if color == "FFFF00":
                series.graphicalProperties.line.dashStyle = "sysDot"  # Dotted
                series.graphicalProperties.line.width = pixels_to_EMU(2)
    
    
    def get_current_row_ranges(self):
        current_col = 2  # assuming Load Current is column 2
        last_row = self.ws.max_row
        current_values = {}
        start_row = None
        last_current = None

        for row in range(10, last_row + 1):  # your data start row is 10
            current = self.ws.cell(row=row, column=current_col).value
            if current != last_current:
                if last_current is not None:
                    current_values[last_current] = (start_row, row - 1)
                start_row = row
                last_current = current

        # Add last current range
        if last_current is not None:
            current_values[last_current] = (start_row, last_row)

        return current_values


    def execute_plot(self):
        if self.keyword not in ["Voltage"]:
            return

        data_start_row = 9  # starting data row
        test_num_col     = 9
        voltage_col = 9
        readback_col = 11
        load_current_col = 7
        psu_voltage_col = 6
        upper_limit_col = 13
        lower_limit_col = 14

        last_row = self.ws.max_row
        last_col = self.ws.max_column

        # Group rows by Load Current value
        current_groups = {}
        current_value = None
        start_row = data_start_row

        for row in range(data_start_row, last_row + 1):
            cell_value = self.ws.cell(row=row, column=load_current_col).value
            if cell_value != current_value:
                if current_value is not None:
                    current_groups[current_value] = (start_row, row - 1)
                current_value = cell_value
                start_row = row
        if current_value is not None:
            current_groups[current_value] = (start_row, last_row)

        # Create the chart Voltage Dependency
        chart = LineChart()
        chart.title = f"{self.keyword} Programming Accuracy"
        chart.y_axis.title = "Voltage"
        chart.x_axis.title = "PSU Voltage Set"
        chart.legend.position = "b"

        for current, (start, end) in current_groups.items():
            x_axis = Reference(self.ws, min_col=psu_voltage_col, min_row=start, max_row=end)
            
            # Voltage series
            values = Reference(self.ws, min_col=voltage_col, min_row=start, max_row=end)
            series = Series(values, title=f"Voltage (Current {current})")
            chart.series.append(series)
            
            # Upper limit series
            upper_values = Reference(self.ws, min_col=upper_limit_col, min_row=start, max_row=end)
            upper_series = Series(upper_values, title="Upper Limit")
            chart.series.append(upper_series)
            
            # Lower limit series
            lower_values = Reference(self.ws, min_col=lower_limit_col, min_row=start, max_row=end)
            lower_series = Series(lower_values, title="Lower Limit")
            chart.series.append(lower_series)
            
            chart.set_categories(x_axis)

        # Optional: style limit lines differently
        for s in chart.series[-2:]:  # last two are upper/lower limit
            s.graphicalProperties.line.solidFill = "FF0000"  # red
            s.graphicalProperties.line.dashStyle = "dash"
            s.graphicalProperties.line.width = pixels_to_EMU(2)

        # Add chart to worksheet
        self.ws.add_chart(chart, "T9")
        # Second chart: Readback Voltage
        chart2 = LineChart()
        chart2.title = "Readback Voltage"
        chart2.y_axis.title = "Voltage"
        chart2.x_axis.title = "PSU Voltage Set"
        chart2.legend.position = "b"

        for current, (start, end) in current_groups.items():
            x_axis = Reference(self.ws, min_col=psu_voltage_col, min_row=start, max_row=end)
            values = Reference(self.ws, min_col=readback_col, min_row=start, max_row=end)
            series = Series(values, title=f"Readback (Current {current})")
            chart2.series.append(series)
            chart2.set_categories(x_axis)

        # Optional: add same upper/lower limit lines for readback
        for current, (start, end) in current_groups.items():
            upper_series = Series(Reference(self.ws, min_col=upper_limit_col, min_row=start, max_row=end), title="Upper Limit")
            lower_series = Series(Reference(self.ws, min_col=lower_limit_col, min_row=start, max_row=end), title="Lower Limit")
            chart2.series.append(upper_series)
            chart2.series.append(lower_series)

        for s in chart2.series[-2:]:
            s.graphicalProperties.line.solidFill = "FF0000"
            s.graphicalProperties.line.dashStyle = "dash"
            s.graphicalProperties.line.width = pixels_to_EMU(2)

        self.ws.add_chart(chart2, "T30")  # position second chart below the first

  
        # ---- Find last data row ----
        voltage_col = 11
        upper_limit_col_raw = 13      # Raw upper limit column (original sheet)
        lower_limit_col_raw = 14      # Raw lower limit column (original sheet)
        
        last_row = self.ws.max_row
        last_col = self.ws.max_column
        
        # ---- Get unique PSU Voltage values ----
        psu_values = sorted({
            self.ws.cell(row=r, column=psu_voltage_col).value
            for r in range(data_start_row, last_row + 1)
        })
        
        # ---- Get unique Current values ----
        current_values = sorted({
            self.ws.cell(row=r, column=load_current_col).value
            for r in range(data_start_row, last_row + 1)
        })
        #--Get unique Upper and Lower limit values--
        upper_limit_values = sorted({
            self.ws.cell(row=r, column=upper_limit_col_raw).value
            for r in range(data_start_row, last_row + 1)
        })
        lower_limit_values = sorted({
            self.ws.cell(row=r, column=lower_limit_col_raw).value
            for r in range(data_start_row, last_row + 1)
        })
        
        # ---- Start grouped table after existing columns ----
        group_start_col = last_col + 3  
        
        # ---- Write header row ----
        self.ws.cell(row=1, column=group_start_col).value = "Load Current"
        
        for i, v in enumerate(psu_values):
            self.ws.cell(row=1, column=group_start_col + 1 + i).value = f"{v} V"
        
        # ---- New columns for aligned spec limits ----
        for i, v in enumerate(upper_limit_values):
            self.ws.cell(row=1, column=group_start_col + 1 + len(psu_values) + i).value = f"{v} V"
        
        for i, v in enumerate(lower_limit_values):
            self.ws.cell(row=1, column=group_start_col + 1 + len(psu_values) + len(upper_limit_values) + i).value = f"{v} V"
        
        # ---- Fill grouped table ----
        for row_idx, current in enumerate(current_values, start=2):
        
            # Write current value
            self.ws.cell(row=row_idx, column=group_start_col).value = current
        
            # Fill each PSU voltage column
            for col_idx, psu in enumerate(psu_values):
                for r in range(data_start_row, last_row + 1):
                    if (self.ws.cell(row=r, column=load_current_col).value == current and
                        self.ws.cell(row=r, column=psu_voltage_col).value == psu):
        
                        measured = self.ws.cell(row=r, column=voltage_col).value
                        self.ws.cell(row=row_idx,
                                    column=group_start_col + 1 + col_idx).value = measured
                        break
            
            for col_idx, v in enumerate(upper_limit_values):
                for r in range(data_start_row, last_row + 1):
                    if (self.ws.cell(row=r, column=load_current_col).value == current and
                        self.ws.cell(row=r, column=upper_limit_col_raw).value == v):
        
                        upper_limit = self.ws.cell(row=r, column=upper_limit_col_raw).value
                        self.ws.cell(row=row_idx,
                                    column=group_start_col + 1 + len(psu_values) + col_idx).value = upper_limit
                        break
            for col_idx, v in enumerate(lower_limit_values):
                for r in range(data_start_row, last_row + 1):
                    if (self.ws.cell(row=r, column=load_current_col).value == current and
                        self.ws.cell(row=r, column=lower_limit_col_raw).value == v):
        
                        lower_limit = self.ws.cell(row=r, column=lower_limit_col_raw).value
                        self.ws.cell(row=row_idx,
                                    column=group_start_col + 1 + len(psu_values) + len(upper_limit_values) + col_idx).value = lower_limit
                        break   
        
        # ---- Chart Creation ----
        chart = LineChart()
        chart.title = "Voltage vs Current (Current Dependency)"
        chart.x_axis.title = "Load Current (A)"
        chart.y_axis.title = "Measured Voltage (V)"
        chart.legend.position = "r"
        
        min_row = 2
        max_row = 1 + len(current_values)
        
        # X-axis reference
        x_ref = Reference(
            self.ws,
            min_col=group_start_col,
            min_row=min_row,
            max_row=max_row
        )
        
        # ---- Add Y series for each PSU ----
        for i, v in enumerate(psu_values):
            y_ref = Reference(
                self.ws,
                min_col=group_start_col + 1 + i,
                min_row=min_row,
                max_row=max_row
            )
        
            series = Series(y_ref, title=f"{v} V")
            series.xvalues = x_ref
            chart.series.append(series)

       
        
        # Styling (red dashed lines)
        for s in (upper_series, lower_series):
            s.graphicalProperties.line.solidFill = "FF0000"
            s.graphicalProperties.line.dashStyle = "dash"
            s.graphicalProperties.line.width = 20000
        
        chart.series.append(upper_series)
        chart.series.append(lower_series)
        
        # Add chart to sheet
        self.ws.add_chart(chart, "T50")
        
        print("✅ Current Dependency Chart Generated (with aligned limits)")
