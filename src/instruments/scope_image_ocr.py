"""Batch OCR and Excel export for Keysight oscilloscope screenshots."""

import base64
import binascii
import json
import re
import subprocess
import tempfile
import threading
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import (
    QDialog,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)


SLOT_COUNT = 8
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp"}
EXCLUDED_DIRECTORIES = {
    "scope_ocr_analysis",
    "waveform_anomaly_analysis",
    "abnormal_waveforms",
}

WINDOWS_OCR_SCRIPT = r"""
$ErrorActionPreference = "Stop"
Add-Type -AssemblyName System.Runtime.WindowsRuntime

function Await-Result($operation, [Type]$resultType) {
    $method = [System.WindowsRuntimeSystemExtensions].GetMethods() |
        Where-Object {
            $_.Name -eq "AsTask" -and
            $_.IsGenericMethod -and
            $_.GetParameters().Count -eq 1
        } |
        Select-Object -First 1
    $task = $method.MakeGenericMethod($resultType).Invoke($null, @($operation))
    $task.Wait()
    return $task.Result
}

function Write-EncodedJson($value) {
    $json = $value | ConvertTo-Json -Compress
    $bytes = [System.Text.Encoding]::UTF8.GetBytes($json)
    [Console]::Out.WriteLine([Convert]::ToBase64String($bytes))
}

[void][Windows.Storage.StorageFile, Windows.Storage, ContentType=WindowsRuntime]
[void][Windows.Storage.Streams.IRandomAccessStream, Windows.Storage.Streams, ContentType=WindowsRuntime]
[void][Windows.Graphics.Imaging.BitmapDecoder, Windows.Graphics.Imaging, ContentType=WindowsRuntime]
[void][Windows.Graphics.Imaging.SoftwareBitmap, Windows.Graphics.Imaging, ContentType=WindowsRuntime]
[void][Windows.Media.Ocr.OcrEngine, Windows.Foundation, ContentType=WindowsRuntime]
[void][Windows.Media.Ocr.OcrResult, Windows.Foundation, ContentType=WindowsRuntime]

$engine = [Windows.Media.Ocr.OcrEngine]::TryCreateFromUserProfileLanguages()
if ($null -eq $engine) {
    throw "Windows OCR is unavailable for the current user language."
}

$items = [Console]::In.ReadToEnd() | ConvertFrom-Json
foreach ($item in $items) {
    try {
        $file = Await-Result (
            [Windows.Storage.StorageFile]::GetFileFromPathAsync([string]$item.path)
        ) ([Windows.Storage.StorageFile])
        $stream = Await-Result (
            $file.OpenAsync([Windows.Storage.FileAccessMode]::Read)
        ) ([Windows.Storage.Streams.IRandomAccessStream])
        try {
            $decoder = Await-Result (
                [Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($stream)
            ) ([Windows.Graphics.Imaging.BitmapDecoder])
            $bitmap = Await-Result (
                $decoder.GetSoftwareBitmapAsync()
            ) ([Windows.Graphics.Imaging.SoftwareBitmap])
            try {
                $result = Await-Result (
                    $engine.RecognizeAsync($bitmap)
                ) ([Windows.Media.Ocr.OcrResult])
                Write-EncodedJson ([pscustomobject]@{
                    id = [string]$item.id
                    text = [string]$result.Text
                    error = $null
                })
            }
            finally {
                if ($bitmap -is [System.IDisposable]) { $bitmap.Dispose() }
            }
        }
        finally {
            $stream.Dispose()
        }
    }
    catch {
        Write-EncodedJson ([pscustomobject]@{
            id = [string]$item.id
            text = ""
            error = $_.Exception.Message
        })
    }
}
"""


class ScopeOcrStopped(RuntimeError):
    pass


@dataclass(frozen=True)
class ScopeMeasurement:
    slot: int
    measurement_type: str
    channel: int | None
    display_value: str
    numeric_value: float | None
    unit: str
    si_value: float | None
    si_unit: str
    ocr_label: str


@dataclass(frozen=True)
class ScopeImageRecord:
    image_path: Path
    timestamp: str
    filename_parameter_1: str
    filename_parameter_2: str
    measurements: tuple
    status: str


@dataclass(frozen=True)
class ScopeOcrSummary:
    image_count: int
    measurement_count: int
    partial_count: int
    error_count: int
    output_path: Path


def discover_scope_images(folder):
    folder = Path(folder)
    if not folder.is_dir():
        raise ValueError(f"Image folder does not exist: {folder}")
    images = []
    for path in folder.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in IMAGE_EXTENSIONS:
            continue
        if any(part.lower() in EXCLUDED_DIRECTORIES for part in path.parts):
            continue
        images.append(path)
    return tuple(sorted(images, key=lambda path: path.name.lower()))


def parse_scope_filename(path):
    stem = Path(path).stem
    match = re.match(
        r"^(.*?)_(.*?)_(\d{4})_(\d{2})_(\d{2})_(\d{2})_(\d{2})_(\d{2})$",
        stem,
    )
    if match is None:
        return "", "", ""
    first, second, year, month, day, hour, minute, second_value = match.groups()
    try:
        timestamp = datetime(
            int(year),
            int(month),
            int(day),
            int(hour),
            int(minute),
            int(second_value),
        ).isoformat(sep=" ")
    except ValueError:
        timestamp = ""
    return first, second, timestamp


def _scope_panel(image):
    height, width = image.shape[:2]
    top = int(height * 0.14)
    bottom = int(height * 0.91)
    left = int(width * 0.81)
    if bottom <= top or left >= width:
        raise ValueError("Image is too small for the Keysight measurement panel")
    return image[top:bottom, left:width]


def _channel_from_color(label_region):
    hsv = cv2.cvtColor(label_region, cv2.COLOR_BGR2HSV)
    saturation = hsv[:, :, 1] > 110
    brightness = hsv[:, :, 2] > 100
    hue = hsv[:, :, 0]
    masks = {
        1: saturation & brightness & (hue >= 18) & (hue <= 38),
        2: saturation & brightness & (hue >= 40) & (hue <= 88),
        3: saturation & brightness & (hue >= 90) & (hue <= 132),
        4: saturation & brightness & (hue >= 135) & (hue <= 178),
    }
    counts = {channel: int(mask.sum()) for channel, mask in masks.items()}
    channel, count = max(counts.items(), key=lambda item: item[1])
    return channel if count >= 3 else None


def prepare_scope_ocr_crops(image_path, output_directory, image_index):
    image = cv2.imread(str(image_path), cv2.IMREAD_COLOR)
    if image is None:
        raise ValueError(f"Unable to read image: {image_path}")
    panel = _scope_panel(image)
    panel_height, panel_width = panel.shape[:2]
    gray = cv2.cvtColor(panel, cv2.COLOR_BGR2GRAY)

    crops = []
    channels = {}
    for slot in range(SLOT_COUNT):
        channel_top = int(panel_height * (0.075 + slot * 0.0825))
        channel_height = max(8, int(panel_height * 0.037))
        channel_region = panel[
            channel_top : min(panel_height, channel_top + channel_height),
            int(panel_width * 0.08) : int(panel_width * 0.78),
        ]
        channels[slot + 1] = _channel_from_color(channel_region)

        label_top = int(panel_height * (0.055 + slot * 0.0825))
        label_height = max(10, int(panel_height * 0.055))
        label_region = gray[
            label_top : min(panel_height, label_top + label_height),
            0 : int(panel_width * 0.95),
        ]
        label_crop = cv2.resize(
            label_region,
            None,
            fx=3,
            fy=3,
            interpolation=cv2.INTER_CUBIC,
        )

        tight_value_top = int(panel_height * (0.116 + slot * 0.0825))
        tight_value_height = max(8, int(panel_height * 0.037))
        tight_value_region = gray[
            tight_value_top : min(
                panel_height,
                tight_value_top + tight_value_height,
            ),
            int(panel_width * 0.45) : panel_width,
        ]
        tight_value_crop = cv2.resize(
            tight_value_region,
            None,
            fx=4,
            fy=4,
            interpolation=cv2.INTER_CUBIC,
        )
        tight_value_crop = cv2.threshold(
            tight_value_crop,
            105,
            255,
            cv2.THRESH_BINARY,
        )[1]

        wide_value_top = int(panel_height * (0.100 + slot * 0.0825))
        wide_value_height = max(12, int(panel_height * 0.055))
        wide_value_region = panel[
            wide_value_top : min(
                panel_height,
                wide_value_top + wide_value_height,
            ),
            int(panel_width * 0.25) : panel_width,
        ]
        wide_value_crop = cv2.resize(
            wide_value_region,
            None,
            fx=3,
            fy=3,
            interpolation=cv2.INTER_CUBIC,
        )

        for kind, crop in (
            ("label", label_crop),
            ("value", tight_value_crop),
            ("value_wide", wide_value_crop),
        ):
            border_value = (0, 0, 0) if crop.ndim == 3 else 0
            crop = cv2.copyMakeBorder(
                crop,
                20,
                20,
                25,
                25,
                cv2.BORDER_CONSTANT,
                value=border_value,
            )
            crop_id = f"{image_index}:{slot + 1}:{kind}"
            crop_path = output_directory / f"{image_index}_{slot + 1}_{kind}.png"
            if not cv2.imwrite(str(crop_path), crop):
                raise RuntimeError(f"Unable to write OCR crop: {crop_path}")
            crops.append({"id": crop_id, "path": str(crop_path.resolve())})
    return crops, channels


def run_windows_ocr(items, stop_requested=None):
    if not items:
        return {}
    process = subprocess.Popen(
        [
            "powershell.exe",
            "-NoProfile",
            "-NonInteractive",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            WINDOWS_OCR_SCRIPT,
        ],
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    payload = json.dumps(items)
    stdout, stderr = process.communicate(payload)
    if stop_requested is not None and stop_requested():
        raise ScopeOcrStopped("Scope image OCR stopped by operator")
    if process.returncode:
        raise RuntimeError(
            "Windows OCR failed: " + (stderr.strip() or "unknown PowerShell error")
        )
    results = {}
    for line in stdout.splitlines():
        line = line.strip()
        if not line:
            continue
        result = decode_ocr_result_line(line)
        results[result["id"]] = result
    return results


def decode_ocr_result_line(line):
    try:
        encoded = base64.b64decode(str(line).strip(), validate=True)
        payload = encoded.decode("utf-8")
        result = json.loads(payload)
    except (binascii.Error, UnicodeDecodeError, json.JSONDecodeError) as exception:
        raise RuntimeError(
            "Windows OCR returned an invalid encoded result"
        ) from exception
    if not isinstance(result, dict) or "id" not in result:
        raise RuntimeError("Windows OCR result is missing its item identifier")
    return result


def normalize_measurement_type(label):
    compact = re.sub(r"\s+", " ", str(label or "")).strip()
    lowered = compact.lower()
    if "rms" in lowered:
        return "DC RMS Cycle"
    if "pk" in lowered or "peak" in lowered:
        return "Peak-to-Peak"
    if "freq" in lowered or lowered.startswith("fre"):
        return "Frequency"
    if "phas" in lowered:
        return "Phase"
    return compact


def parse_display_value(text):
    display = re.sub(r"\s+", " ", str(text or "")).strip()
    unit_pattern = (
        r"MHz|kHz|Hz|mV|uV|ÂµV|nV|V|mA|uA|ÂµA|A|ms|us|Âµs|ns|s|Â°|deg"
    )

    def normalize_numeric_token(match):
        number = match.group("number").translate(
            str.maketrans({"I": "1", "l": "1", "O": "0", "o": "0"})
        )
        return f"{number}{match.group('unit')}"

    display = re.sub(
        rf"(?P<number>[-+]?[0-9IlOo]+(?:[.,\s][0-9IlOo]+)*)\s*(?P<unit>{unit_pattern})\b",
        normalize_numeric_token,
        display,
        flags=re.IGNORECASE,
    )
    display = re.sub(r"(?<=\.)\s+(?=\d)", "", display)
    display = re.sub(
        r"(?<=\.)\s*[Il](?=\s*(?:[kMmun]?Hz|[munk]?V|[mun]?A))",
        "1",
        display,
    )
    display = re.sub(r"(?<=\d)\s+[Il](?=\s*(?:[kMmun]?Hz|[munk]?V))", "1", display)
    display = re.sub(
        rf"(?<=\d)\s+(?=\d{{1,3}}\s*(?:{unit_pattern})\b)",
        ".",
        display,
        flags=re.IGNORECASE,
    )
    display = display.replace(",", ".")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", display)
    if match is None:
        return display, None, "", None, ""
    numeric = float(match.group(0))
    unit_match = re.search(
        r"(MHz|kHz|Hz|mV|uV|µV|nV|V|mA|uA|µA|A|ms|us|µs|ns|s|°|deg)",
        display,
        re.IGNORECASE,
    )
    unit = unit_match.group(1) if unit_match else ""
    normalized_unit = unit.replace("µ", "u")
    conversions = {
        "mhz": (1_000_000.0, "Hz"),
        "khz": (1_000.0, "Hz"),
        "hz": (1.0, "Hz"),
        "mv": (0.001, "V"),
        "uv": (0.000001, "V"),
        "nv": (0.000000001, "V"),
        "v": (1.0, "V"),
        "ma": (0.001, "A"),
        "ua": (0.000001, "A"),
        "a": (1.0, "A"),
        "ms": (0.001, "s"),
        "us": (0.000001, "s"),
        "ns": (0.000000001, "s"),
        "s": (1.0, "s"),
        "°": (1.0, "degree"),
        "deg": (1.0, "degree"),
    }
    factor, si_unit = conversions.get(normalized_unit.lower(), (1.0, unit))
    return display, numeric, unit, numeric * factor, si_unit


def _value_candidate(text):
    parsed = parse_display_value(text)
    display, numeric, unit, _si_value, _si_unit = parsed
    score = 0
    if numeric is not None:
        score += 10
    if unit:
        score += 8
    if re.search(r"\d[.,]\d", display):
        score += 2
    if re.search(r"[A-Za-zÂµ°]", display):
        score += 1
    return score, parsed


def _best_display_value(*texts):
    candidates = [_value_candidate(text) for text in texts if str(text or "").strip()]
    if not candidates:
        return "", None, "", None, ""
    return max(candidates, key=lambda candidate: candidate[0])[1]


def _channel_from_label(label):
    match = re.search(r"\(\s*([1-4])", str(label or ""))
    return int(match.group(1)) if match else None


def build_scope_record(image_path, image_index, channels, ocr_results):
    measurements = []
    for slot in range(1, SLOT_COUNT + 1):
        label_result = ocr_results.get(f"{image_index}:{slot}:label", {})
        value_result = ocr_results.get(f"{image_index}:{slot}:value", {})
        wide_value_result = ocr_results.get(
            f"{image_index}:{slot}:value_wide",
            {},
        )
        label = str(label_result.get("text") or "").strip()
        value_text = str(value_result.get("text") or "").strip()
        wide_value_text = str(wide_value_result.get("text") or "").strip()
        if not label and not value_text and not wide_value_text:
            continue
        measurement_type = normalize_measurement_type(label)
        display, numeric, unit, si_value, si_unit = _best_display_value(
            value_text,
            wide_value_text,
        )
        if not measurement_type and si_unit == "Hz":
            measurement_type = "Frequency"
        if (
            measurement_type in {"DC RMS Cycle", "Peak-to-Peak"}
            and numeric is not None
            and not unit
            and re.search(r"\d\s*[MN]$", display, re.IGNORECASE)
        ):
            display = re.sub(r"[MN]$", "V", display, flags=re.IGNORECASE)
            unit = "V"
            si_value = numeric
            si_unit = "V"
        if measurement_type == "Phase" and numeric is not None and not unit:
            unit = "°"
            si_value = numeric
            si_unit = "degree"
        number_match = re.search(r"[-+]?\d+(?:\.\d+)?", display)
        if number_match is not None and unit:
            display = f"{number_match.group(0)}{unit}"
        measurements.append(
            ScopeMeasurement(
                slot=slot,
                measurement_type=measurement_type,
                channel=channels.get(slot) or _channel_from_label(label),
                display_value=display,
                numeric_value=numeric,
                unit=unit,
                si_value=si_value,
                si_unit=si_unit,
                ocr_label=label,
            )
        )
    first, second, timestamp = parse_scope_filename(image_path)
    numeric_count = sum(item.numeric_value is not None for item in measurements)
    complete_count = sum(
        item.numeric_value is not None
        and bool(item.measurement_type)
        and item.channel is not None
        and bool(item.unit)
        for item in measurements
    )
    if numeric_count == 0:
        status = "NO DATA"
    elif complete_count < len(measurements) or complete_count < 2:
        status = "PARTIAL"
    else:
        status = "OK"
    return ScopeImageRecord(
        image_path=Path(image_path),
        timestamp=timestamp,
        filename_parameter_1=first,
        filename_parameter_2=second,
        measurements=tuple(measurements),
        status=status,
    )


def _numeric_filename_parameter(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _record_sort_key(record):
    voltage = _numeric_filename_parameter(record.filename_parameter_1)
    current = _numeric_filename_parameter(record.filename_parameter_2)
    return (
        voltage is None,
        voltage if voltage is not None else float("inf"),
        current is None,
        current if current is not None else float("inf"),
        record.timestamp,
        record.image_path.name.lower(),
    )


def assign_result_numbers(records):
    """Number repeated captures within each voltage/current condition."""
    result_numbers = {}
    condition_counts = {}
    for record in sorted(records, key=_record_sort_key):
        condition = (
            record.filename_parameter_1,
            record.filename_parameter_2,
        )
        result_number = condition_counts.get(condition, 0) + 1
        condition_counts[condition] = result_number
        result_numbers[record.image_path] = result_number
    return result_numbers


def _add_power_relationship_chart(
    charts_sheet,
    data_sheets,
    value_column,
    title,
    y_axis_title,
    anchor,
    color,
):
    chart = ScatterChart()
    chart.title = title
    chart.style = 13
    chart.scatterStyle = "marker"
    chart.height = 9
    chart.width = 16
    chart.x_axis.title = "Power (W)"
    chart.y_axis.title = y_axis_title
    chart.x_axis.numFmt = "0.00"
    chart.y_axis.numFmt = "0.000"
    data_sheets = tuple(data_sheets)
    result_colors = (color, "ED7D31", "70AD47", "A5A5A5", "7030A0")
    if len(data_sheets) == 1:
        chart.legend = None
    for index, data_sheet in enumerate(data_sheets):
        power_values = Reference(
            data_sheet,
            min_col=4,
            min_row=2,
            max_row=data_sheet.max_row,
        )
        measurement_values = Reference(
            data_sheet,
            min_col=value_column,
            min_row=2,
            max_row=data_sheet.max_row,
        )
        series_title = (
            y_axis_title if len(data_sheets) == 1 else data_sheet.title
        )
        series = Series(measurement_values, power_values, title=series_title)
        series_color = result_colors[index % len(result_colors)]
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = series_color
        series.marker.graphicalProperties.line.solidFill = series_color
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
    charts_sheet.add_chart(chart, anchor)


def export_scope_records(records, output_path, source_folder):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    result_numbers = assign_result_numbers(records)
    records = tuple(sorted(
        records,
        key=lambda record: (
            result_numbers[record.image_path],
            *_record_sort_key(record),
        ),
    ))
    maximum_result = max(result_numbers.values(), default=1)
    workbook = Workbook()
    images_sheet = workbook.active
    images_sheet.title = "Images"
    summary_sheet = workbook.create_sheet("Summary")
    measurement_sheet = workbook.create_sheet("Measurements")
    charts_sheet = workbook.create_sheet("Charts")
    guide_sheet = workbook.create_sheet("Guide")

    images_headers = (
        "Image",
        "Full Path",
        "Timestamp",
        "Voltage",
        "Current",
        "Result",
        "Measurement Count",
        "Status",
    )
    measurement_headers = (
        "Image",
        "Timestamp",
        "Voltage",
        "Current",
        "Result",
        "Slot",
        "Measurement Type",
        "Channel",
        "Displayed Value",
        "Numeric Value",
        "Displayed Unit",
        "SI Value",
        "SI Unit",
        "OCR Label",
        "Image Path",
    )
    images_sheet.append(images_headers)
    measurement_sheet.append(measurement_headers)

    summary_headers = (
        "Result",
        "Voltage",
        "Current",
        "Power (W)",
        "DC RMS Cycle",
        "Pk-Pk(4)",
        "Freq(3)",
        "Pk-Pk(3)",
        "DC RMS Cyc(3)",
        "DC RMS Cyc(2)",
        "Pk-Pk(2)",
        "Phase(1->3)",
        "Oscilloscope Image",
        "Noise Pk-Pk(2) (V)",
        "Phase Angle (deg)",
    )
    summary_sheet.append(summary_headers)
    result_sheets = {}
    if maximum_result > 1:
        for result_number in range(1, maximum_result + 1):
            result_sheet = workbook.create_sheet(f"Result {result_number}")
            result_sheet.append(summary_headers)
            result_sheets[result_number] = result_sheet

    for record in records:
        result_number = result_numbers[record.image_path]
        images_sheet.append(
            (
                record.image_path.name,
                str(record.image_path),
                record.timestamp,
                record.filename_parameter_1,
                record.filename_parameter_2,
                result_number,
                len(record.measurements),
                record.status,
            )
        )
        for measurement in record.measurements:
            measurement_sheet.append(
                (
                    record.image_path.name,
                    record.timestamp,
                    record.filename_parameter_1,
                    record.filename_parameter_2,
                    result_number,
                    measurement.slot,
                    measurement.measurement_type,
                    measurement.channel,
                    measurement.display_value,
                    measurement.numeric_value,
                    measurement.unit,
                    measurement.si_value,
                    measurement.si_unit,
                    measurement.ocr_label,
                    str(record.image_path),
                )
            )
        values_by_slot = {
            measurement.slot: measurement.display_value
            for measurement in record.measurements
            if measurement.display_value
        }
        measurements_by_slot = {
            measurement.slot: measurement
            for measurement in record.measurements
        }
        summary_values = (
            result_number,
            _numeric_filename_parameter(record.filename_parameter_1),
            _numeric_filename_parameter(record.filename_parameter_2),
            None,
            *(values_by_slot.get(slot) for slot in range(1, SLOT_COUNT + 1)),
            None,
            getattr(measurements_by_slot.get(7), "si_value", None),
            getattr(measurements_by_slot.get(8), "si_value", None),
        )
        summary_sheet.append(summary_values)
        summary_row = summary_sheet.max_row
        summary_sheet.cell(summary_row, 4).value = f"=B{summary_row}*C{summary_row}"
        result_sheet = result_sheets.get(result_number)
        if result_sheet is not None:
            result_sheet.append(summary_values)
            result_row = result_sheet.max_row
            result_sheet.cell(result_row, 4).value = (
                f"=B{result_row}*C{result_row}"
            )
            result_sheet.cell(result_row, 13).value = str(record.image_path)
        try:
            scope_image = ExcelImage(str(record.image_path))
            scale = min(400 / scope_image.width, 315 / scope_image.height, 1.0)
            scope_image.width = int(scope_image.width * scale)
            scope_image.height = int(scope_image.height * scale)
            summary_sheet.add_image(scope_image, f"M{summary_row}")
            summary_sheet.row_dimensions[summary_row].height = max(
                30,
                scope_image.height * 0.75 + 8,
            )
        except (OSError, ValueError) as exception:
            summary_sheet.cell(summary_row, 13).value = (
                f"Unable to embed image: {exception}"
            )

    charts_sheet["A1"] = "Power Relationship Charts"
    charts_sheet["A1"].font = Font(size=16, bold=True, color="1F4E78")
    charts_sheet.sheet_view.showGridLines = False
    chart_data_sheets = tuple(result_sheets.values()) or (summary_sheet,)
    _add_power_relationship_chart(
        charts_sheet,
        chart_data_sheets,
        14,
        "Power vs Noise Pk-Pk(2)",
        "Noise Pk-Pk(2) (V)",
        "A3",
        "4472C4",
    )
    _add_power_relationship_chart(
        charts_sheet,
        chart_data_sheets,
        15,
        "Power vs Phase Angle",
        "Phase Angle (deg)",
        "A21",
        "ED7D31",
    )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in (
        images_sheet,
        summary_sheet,
        measurement_sheet,
        *result_sheets.values(),
    ):
        sheet.freeze_panes = "A2"
        if sheet is summary_sheet:
            sheet.auto_filter.ref = f"A1:M{sheet.max_row}"
        else:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in zip("ABCDEFGH", (34, 90, 22, 22, 22, 10, 20, 14)):
        images_sheet.column_dimensions[column].width = width
    summary_widths = (10, 12, 12, 12, 20, 18, 18, 18, 20, 20, 18, 20, 58)
    for index, width in enumerate(summary_widths, start=1):
        summary_sheet.column_dimensions[
            summary_sheet.cell(1, index).column_letter
        ].width = width
    for row in summary_sheet.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = "0.############"
    summary_sheet.column_dimensions["N"].hidden = True
    summary_sheet.column_dimensions["O"].hidden = True
    measurement_widths = (34, 22, 22, 22, 10, 8, 22, 10, 20, 16, 16, 16, 12, 28, 90)
    for index, width in enumerate(measurement_widths, start=1):
        measurement_sheet.column_dimensions[
            measurement_sheet.cell(1, index).column_letter
        ].width = width
    for result_sheet in result_sheets.values():
        for index, width in enumerate(summary_widths, start=1):
            result_sheet.column_dimensions[
                result_sheet.cell(1, index).column_letter
            ].width = width
        for row in result_sheet.iter_rows(min_row=2, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = "0.############"
        result_sheet.column_dimensions["N"].hidden = True
        result_sheet.column_dimensions["O"].hidden = True

    guide_rows = (
        ("Scope Screenshot OCR Export", ""),
        ("Source Folder", str(source_folder)),
        ("Generated", datetime.now().isoformat(sep=" ", timespec="seconds")),
        ("Method", "OpenCV fixed-region preprocessing plus Windows OCR"),
        (
            "Measurement Type",
            "Normalized from the label displayed in each measurement slot.",
        ),
        (
            "Summary Sorting",
            "Rows are grouped by Result number, then sorted by numeric filename voltage and current from low to high.",
        ),
        (
            "Result Number",
            "Repeated screenshots with the same filename voltage/current condition are numbered chronologically. Result 1 is the earlier capture and Result 2 is the later capture. Separate Result worksheets contain lightweight data copies without duplicate embedded images.",
        ),
        (
            "Summary Values",
            "The fixed Summary columns preserve the value and unit displayed on the oscilloscope.",
        ),
        (
            "Power",
            "Power is calculated in watts for each row as filename Voltage multiplied by filename Current.",
        ),
        (
            "Charts",
            "The Charts sheet plots Power versus Noise Pk-Pk(2) and Power versus Phase Angle using numeric SI values extracted by OCR.",
        ),
        (
            "Embedded Images",
            "Each Summary row includes a resized copy of its original oscilloscope screenshot.",
        ),
        (
            "Channel",
            "Detected from Keysight channel colors: yellow=1, green=2, blue=3, magenta=4.",
        ),
        (
            "SI Value",
            "Numeric value converted to base V, A, Hz, seconds, or degrees when a unit is recognized.",
        ),
        (
            "PARTIAL",
            "OCR found a label or value that could not be parsed completely; review the original image.",
        ),
        (
            "Important",
            "OCR can misread small characters. Keep the original screenshots and verify critical values.",
        ),
    )
    for row in guide_rows:
        guide_sheet.append(row)
    guide_sheet["A1"].font = Font(size=16, bold=True, color="1F4E78")
    guide_sheet.column_dimensions["A"].width = 24
    guide_sheet.column_dimensions["B"].width = 110
    for row in guide_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output_path)


def compile_scope_image_folder(
    folder,
    output_path=None,
    progress_callback=None,
    record_callback=None,
    stop_requested=None,
    ocr_backend=run_windows_ocr,
):
    folder = Path(folder)
    images = discover_scope_images(folder)
    if not images:
        raise ValueError("No oscilloscope images were found in the selected folder")
    if output_path is None:
        output_path = folder / "scope_ocr_analysis" / "scope_measurements.xlsx"
    output_path = Path(output_path)
    records = []
    errors = 0
    chunk_size = 20

    with tempfile.TemporaryDirectory(prefix="hornbill_scope_ocr_") as temporary:
        temporary_root = Path(temporary)
        for chunk_start in range(0, len(images), chunk_size):
            if stop_requested is not None and stop_requested():
                raise ScopeOcrStopped("Scope image OCR stopped by operator")
            chunk = images[chunk_start : chunk_start + chunk_size]
            crop_items = []
            prepared = []
            for offset, image_path in enumerate(chunk):
                image_index = chunk_start + offset
                try:
                    items, channels = prepare_scope_ocr_crops(
                        image_path,
                        temporary_root,
                        image_index,
                    )
                    crop_items.extend(items)
                    prepared.append((image_index, image_path, channels, None))
                except Exception as exception:
                    prepared.append((image_index, image_path, {}, exception))
            ocr_results = ocr_backend(crop_items, stop_requested=stop_requested)
            for image_index, image_path, channels, preparation_error in prepared:
                if preparation_error is not None:
                    errors += 1
                    record = ScopeImageRecord(
                        image_path=image_path,
                        timestamp="",
                        filename_parameter_1="",
                        filename_parameter_2="",
                        measurements=(),
                        status="ERROR",
                    )
                else:
                    record = build_scope_record(
                        image_path,
                        image_index,
                        channels,
                        ocr_results,
                    )
                records.append(record)
                if record_callback is not None:
                    record_callback(record)
                completed = len(records)
                if progress_callback is not None:
                    progress_callback(
                        completed,
                        len(images),
                        f"OCR {completed}/{len(images)}: {image_path.name} ({record.status})",
                    )

    export_scope_records(records, output_path, folder)
    return ScopeOcrSummary(
        image_count=len(records),
        measurement_count=sum(len(record.measurements) for record in records),
        partial_count=sum(record.status != "OK" for record in records),
        error_count=errors,
        output_path=output_path,
    )


class ScopeImageOcrWorker(QThread):
    progress = pyqtSignal(int)
    message = pyqtSignal(str)
    record_ready = pyqtSignal(object)
    completed = pyqtSignal(object)
    stopped = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, folder, output_path, parent=None):
        super().__init__(parent)
        self.folder = Path(folder)
        self.output_path = Path(output_path)
        self._stop_requested = threading.Event()

    def stop(self):
        self._stop_requested.set()

    def _progress(self, completed, total, message):
        self.progress.emit(int(completed * 100 / total))
        self.message.emit(message)

    def run(self):
        try:
            summary = compile_scope_image_folder(
                self.folder,
                self.output_path,
                progress_callback=self._progress,
                record_callback=self.record_ready.emit,
                stop_requested=self._stop_requested.is_set,
            )
            self.completed.emit(summary)
        except ScopeOcrStopped:
            self.stopped.emit()
        except Exception as exception:
            self.message.emit(f"ERROR: {exception}\n{traceback.format_exc()}")
            self.error.emit(str(exception))


class ScopeImageOcrDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Oscilloscope Image OCR to Excel")
        self.resize(1180, 760)
        self.worker = None
        self._preview_pixmap = None

        title = QLabel("Oscilloscope Screenshot Data Extractor")
        title.setStyleSheet("font-size: 24px; font-weight: 700; color: #17365d;")
        subtitle = QLabel(
            "Scan Keysight oscilloscope screenshots, extract displayed measurement "
            "labels and values, detect channel colors, and compile one Excel file."
        )
        subtitle.setWordWrap(True)

        self.folder = QLineEdit()
        self.folder.setPlaceholderText("Folder containing oscilloscope images")
        folder_button = QPushButton("Browse Images...")
        folder_button.clicked.connect(self._browse_folder)
        folder_layout = QHBoxLayout()
        folder_layout.addWidget(self.folder)
        folder_layout.addWidget(folder_button)

        self.output_path = QLineEdit()
        self.output_path.setPlaceholderText("Output Excel file")
        output_button = QPushButton("Save As...")
        output_button.clicked.connect(self._browse_output)
        output_layout = QHBoxLayout()
        output_layout.addWidget(self.output_path)
        output_layout.addWidget(output_button)

        form = QFormLayout()
        form.addRow("Image Folder:", folder_layout)
        form.addRow("Excel Output:", output_layout)

        self.start_button = QPushButton("Extract and Compile")
        self.stop_button = QPushButton("Stop")
        self.stop_button.setEnabled(False)
        self.start_button.clicked.connect(self._start)
        self.stop_button.clicked.connect(self._stop)
        buttons = QHBoxLayout()
        buttons.addWidget(self.start_button)
        buttons.addWidget(self.stop_button)
        buttons.addStretch(1)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.status = QLabel(
            "Ready. OCR results should be verified against critical screenshots."
        )
        self.status.setWordWrap(True)

        self.results = QTableWidget(0, 3)
        self.results.setHorizontalHeaderLabels(("Image", "Measurements", "Status"))
        self.results.horizontalHeader().setStretchLastSection(True)
        self.results.setColumnWidth(0, 300)
        self.results.setColumnWidth(1, 110)
        self.results.currentCellChanged.connect(self._show_selected)

        self.preview = QLabel("Select an extracted image to preview it")
        self.preview.setAlignment(Qt.AlignCenter)
        self.preview.setMinimumSize(560, 390)
        self.preview.setStyleSheet(
            "background: #08111f; color: #94a3b8; border-radius: 10px;"
        )
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(self.results)
        splitter.addWidget(self.preview)
        splitter.setSizes((430, 720))

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMaximumHeight(140)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(22, 22, 22, 22)
        layout.setSpacing(12)
        layout.addWidget(title)
        layout.addWidget(subtitle)
        layout.addLayout(form)
        layout.addLayout(buttons)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.status)
        layout.addWidget(splitter, stretch=1)
        layout.addWidget(self.log)

    def _browse_folder(self):
        folder = QFileDialog.getExistingDirectory(
            self,
            "Select Oscilloscope Image Folder",
            self.folder.text().strip(),
        )
        if folder:
            self.folder.setText(folder)
            default_output = (
                Path(folder) / "scope_ocr_analysis" / "scope_measurements.xlsx"
            )
            self.output_path.setText(str(default_output))

    def _browse_output(self):
        output, _ = QFileDialog.getSaveFileName(
            self,
            "Save Extracted Scope Data",
            self.output_path.text().strip() or "scope_measurements.xlsx",
            "Excel Workbook (*.xlsx)",
        )
        if output:
            if not output.lower().endswith(".xlsx"):
                output += ".xlsx"
            self.output_path.setText(output)

    def _start(self):
        folder = self.folder.text().strip()
        output = self.output_path.text().strip()
        try:
            images = discover_scope_images(folder)
            if not images:
                raise ValueError("No supported images were found")
            if not output:
                raise ValueError("Select an Excel output file")
        except ValueError as exception:
            QMessageBox.warning(self, "Invalid OCR Settings", str(exception))
            return

        self.results.setRowCount(0)
        self.log.clear()
        self.progress_bar.setValue(0)
        self.status.setText(f"Extracting measurements from {len(images)} image(s)...")
        self.worker = ScopeImageOcrWorker(folder, output, self)
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.message.connect(self.log.append)
        self.worker.record_ready.connect(self._record_ready)
        self.worker.completed.connect(self._completed)
        self.worker.stopped.connect(self._stopped)
        self.worker.error.connect(self._error)
        self.worker.finished.connect(self._finished)
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.worker.start()

    def _record_ready(self, record):
        row = self.results.rowCount()
        self.results.insertRow(row)
        image_item = QTableWidgetItem(record.image_path.name)
        image_item.setData(Qt.UserRole, str(record.image_path))
        self.results.setItem(row, 0, image_item)
        self.results.setItem(row, 1, QTableWidgetItem(str(len(record.measurements))))
        self.results.setItem(row, 2, QTableWidgetItem(record.status))
        if row == 0:
            self.results.selectRow(0)

    def _show_selected(self, row, _column, _previous_row, _previous_column):
        if row < 0:
            return
        item = self.results.item(row, 0)
        if item is None:
            return
        pixmap = QPixmap(item.data(Qt.UserRole))
        if pixmap.isNull():
            self.preview.setText("Unable to display selected image")
            self._preview_pixmap = None
            return
        self._preview_pixmap = pixmap
        self._update_preview()

    def _update_preview(self):
        if self._preview_pixmap is None:
            return
        self.preview.setPixmap(
            self._preview_pixmap.scaled(
                self.preview.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation,
            )
        )

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._update_preview()

    def _stop(self):
        if self.worker is not None:
            self.worker.stop()
            self.stop_button.setEnabled(False)
            self.status.setText("Stop requested; waiting for the current OCR batch...")

    def _completed(self, summary):
        self.progress_bar.setValue(100)
        self.status.setText(
            f"Completed: {summary.image_count} image(s), "
            f"{summary.measurement_count} measurement(s), "
            f"{summary.partial_count} image(s) need review. Excel: "
            f"{summary.output_path}"
        )

    def _stopped(self):
        self.status.setText("OCR extraction stopped by operator.")

    def _error(self, message):
        self.status.setText("OCR extraction failed.")
        QMessageBox.critical(self, "Oscilloscope OCR Error", message)

    def _finished(self):
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.worker = None

    def closeEvent(self, event):
        if self.worker is not None and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait(5000)
        super().closeEvent(event)
