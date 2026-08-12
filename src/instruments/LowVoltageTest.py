"""Hornbill low-voltage DIAG sweep with oscilloscope screenshots."""

import csv
import time
import threading
import traceback
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import (
    QComboBox,
    QCheckBox,
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSpinBox,
    QTextEdit,
    QVBoxLayout,
)

from SCPI_Library.Keysight import DMM_344XXA, DMM_3458A, Hornbill, Oscilloscope


LOW_VOLTAGE_INCREMENT = 100
DEFAULT_INITIAL_POINT = 800
DEFAULT_FINAL_POINT = 6000


class LowVoltageTestStopped(RuntimeError):
    pass


class LowVoltageExcelReport:
    HEADERS = (
        "Loop",
        "DUT Channel",
        "DIAG Point",
        "DMM Voltage (V)",
        "Scope VMIN (V)",
        "Scope VMAX (V)",
        "Scope VPP (mV)",
        "Screenshot File",
        "Waveform",
    )

    def __init__(self, path, configuration):
        self.path = Path(path)
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Low Voltage Capture"
        self.worksheet.append(self.HEADERS)
        self.worksheet.freeze_panes = "A2"
        self.worksheet.auto_filter.ref = "A1:I1"
        self.worksheet.row_dimensions[1].height = 24
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in self.worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        widths = (10, 14, 14, 18, 16, 16, 16, 48, 58)
        for column, width in zip("ABCDEFGHI", widths):
            self.worksheet.column_dimensions[column].width = width

        metadata = self.workbook.create_sheet("Test Settings")
        metadata.append(("Setting", "Value"))
        metadata["A1"].font = Font(bold=True, color="FFFFFF")
        metadata["B1"].font = Font(bold=True, color="FFFFFF")
        metadata["A1"].fill = header_fill
        metadata["B1"].fill = header_fill
        settings = (
            ("Hornbill Address", configuration.get("PSU", "")),
            ("Oscilloscope Address", configuration.get("OSC", "")),
            ("DMM Enabled", bool(configuration.get("DMM_Enabled", False))),
            ("DMM Model", configuration.get("DMM_Model", "")),
            ("Initial DIAG Point", configuration.get("LowVoltageInitialPoint", "")),
            ("Final DIAG Point", configuration.get("LowVoltageFinalPoint", "")),
            ("DIAG Increment", configuration.get("LowVoltageIncrement", "")),
            ("DIAG Settling Delay (s)", configuration.get("updatedelay", "")),
            ("Scope Capture Delay (s)", configuration.get("ScopeRunCaptureDelay", "")),
            ("Scope VPP Enabled", bool(configuration.get("ScopeVppEnabled", False))),
            ("Ink Saver", configuration.get("InkSaver", "")),
        )
        for setting in settings:
            metadata.append(setting)
        metadata.column_dimensions["A"].width = 28
        metadata.column_dimensions["B"].width = 70

    def append(self, record, image_path):
        self.worksheet.append(
            tuple(record.get(header) for header in self.HEADERS[:-1]) + (None,)
        )
        row = self.worksheet.max_row
        self.worksheet.row_dimensions[row].height = 195
        for cell in self.worksheet[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        image = ExcelImage(str(image_path))
        image.width = 400
        image.height = 250
        self.worksheet.add_image(image, f"I{row}")

    def save(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.path)


def low_voltage_points(
    initial_point=DEFAULT_INITIAL_POINT,
    final_point=DEFAULT_FINAL_POINT,
    increment=LOW_VOLTAGE_INCREMENT,
):
    initial_point = int(initial_point)
    final_point = int(final_point)
    increment = int(increment)
    if increment <= 0:
        raise ValueError("Low-voltage increment must be greater than zero")
    if initial_point > final_point:
        raise ValueError(
            "Low-voltage initial point cannot exceed the final point"
        )
    if (final_point - initial_point) % increment:
        raise ValueError(
            "Low-voltage final point must align with the selected increment"
        )
    return tuple(range(initial_point, final_point + 1, increment))


def diag_channel_selector(channel):
    channel = int(channel)
    if not 1 <= channel <= 4:
        raise ValueError("Hornbill channel must be between 1 and 4")
    return channel - 1


def strip_ieee_binary_block(data):
    image_data = bytes(data)
    if not image_data.startswith(b"#"):
        return image_data
    if len(image_data) < 2 or not image_data[1:2].isdigit():
        raise RuntimeError("Oscilloscope returned an invalid binary header")
    length_digits = int(image_data[1:2])
    header_end = 2 + length_digits
    if len(image_data) < header_end:
        raise RuntimeError("Oscilloscope returned an incomplete binary header")
    try:
        payload_length = int(image_data[2:header_end])
    except ValueError as exception:
        raise RuntimeError(
            "Oscilloscope returned an invalid payload length"
        ) from exception
    payload_end = header_end + payload_length
    if len(image_data) < payload_end:
        raise RuntimeError("Oscilloscope screenshot payload is incomplete")
    return image_data[header_end:payload_end]


class LowVoltageScopeCaptureTest:
    def __init__(
        self,
        hornbill_factory=Hornbill,
        oscilloscope_factory=Oscilloscope,
        dmm_344xxa_factory=DMM_344XXA,
        dmm_3458a_factory=DMM_3458A,
    ):
        self.hornbill_factory = hornbill_factory
        self.oscilloscope_factory = oscilloscope_factory
        self.dmm_factories = {
            "344xxA": dmm_344xxa_factory,
            "3458A": dmm_3458a_factory,
        }

    @staticmethod
    def _channels(value):
        if isinstance(value, (range, list, tuple, set)):
            return tuple(int(channel) for channel in value)
        return (int(value),)

    @staticmethod
    def _checkpoint(worker):
        if worker is not None:
            worker.checkpoint()

    @staticmethod
    def _wait(worker, seconds):
        if worker is not None:
            worker.interruptible_sleep(seconds)
        else:
            time.sleep(max(0.0, float(seconds)))

    @staticmethod
    def _emit(worker, message):
        if worker is not None:
            worker.progress.emit(message)

    @staticmethod
    def _check_dut_error(dut, point):
        response = str(dut.instr.query("SYST:ERR?")).strip()
        if not (response.startswith("0") or response.startswith("+0")):
            raise RuntimeError(
                f"Hornbill rejected low-voltage point {point}: {response}"
            )

    @staticmethod
    def _configure_scope(scope, configuration):
        channel = int(configuration["OSC_Channel"])
        scope.instr.write(f":CHANnel{channel}:DISPlay ON")
        scope.hardcopy(configuration.get("InkSaver", "ON"))
        optional_commands = (
            (
                configuration.get("Channel_CouplingMode"),
                lambda value: f":CHANnel{channel}:COUPling {value}",
            ),
            (
                configuration.get("TimeScale"),
                lambda value: f":TIMebase:RANGe {value}",
            ),
            (
                configuration.get("VerticalScale"),
                lambda value: f":CHANnel{channel}:SCALe {value}",
            ),
            (
                configuration.get("Trigger_SweepMode"),
                lambda value: f":TRIGger:SWEep {value}",
            ),
        )
        for value, command_factory in optional_commands:
            if value not in (None, ""):
                scope.instr.write(command_factory(value))

    def _create_dmm(self, configuration):
        if not configuration.get("DMM_Enabled", False):
            return None
        model = str(configuration.get("DMM_Model", "344xxA"))
        try:
            factory = self.dmm_factories[model]
        except KeyError as exception:
            raise ValueError(f"Unsupported DMM model: {model}") from exception
        dmm = factory(configuration["DMM"])
        try:
            if model == "3458A":
                dmm.setDCV("10")
                dmm.setNPLC("10")
                dmm.setAutoZeroMode("ON")
        except Exception:
            session = getattr(dmm, "instr", None)
            if session is not None:
                session.close()
            raise
        return dmm

    @staticmethod
    def _measure_dmm(dmm, configuration):
        if dmm is None:
            return None
        if configuration.get("DMM_Model") == "3458A":
            return float(dmm.queryMeasurement())
        return float(dmm.instr.query("MEAS:VOLT:DC?"))

    @staticmethod
    def _first_numeric_value(value):
        if isinstance(value, (list, tuple)):
            if not value:
                raise RuntimeError("Oscilloscope returned no measurement value")
            value = value[0]
        return float(value)

    @classmethod
    def _measure_scope_voltage(cls, scope, configuration):
        if not configuration.get("ScopeVppEnabled", False):
            return None, None, None
        source = f"CHANNEL{int(configuration['OSC_Channel'])}"
        minimum = cls._first_numeric_value(scope.measureVMIN(source))
        maximum = cls._first_numeric_value(scope.measureVMAX(source))
        return minimum, maximum, (maximum - minimum) * 1000.0

    @staticmethod
    def _output_directory(configuration, worker):
        if worker is not None and worker.run_context is not None:
            root = worker.run_context.storage.charts
        else:
            root = Path(
                configuration.get("rawdir")
                or configuration.get("savedir")
                or configuration.get("savelocation")
            )
        output_directory = Path(root) / "low_voltage_scope"
        output_directory.mkdir(parents=True, exist_ok=True)
        return output_directory

    def execute(self, configuration, loop_index=0, worker=None):
        points = low_voltage_points(
            configuration["LowVoltageInitialPoint"],
            configuration["LowVoltageFinalPoint"],
            configuration.get("LowVoltageIncrement", LOW_VOLTAGE_INCREMENT),
        )
        channels = self._channels(configuration["PSU_Channel"])
        output_directory = self._output_directory(configuration, worker)
        delay = float(configuration.get("updatedelay") or 0)
        scope_run_delay = float(
            configuration.get("ScopeRunCaptureDelay", 3.0)
        )
        dut = None
        scope = None
        dmm = None
        measurement_file = None
        measurement_writer = None
        excel_report = None
        enabled_channels = []
        total = len(points) * len(channels)
        completed = 0
        captured_images = []

        try:
            dut = self.hornbill_factory(configuration["PSU"])
            scope = self.oscilloscope_factory(configuration["OSC"])
            dmm = self._create_dmm(configuration)
            self._configure_scope(scope, configuration)
            if configuration.get("ExcelReportEnabled", False):
                report_path = output_directory / (
                    f"low_voltage_scope_report_loop_{loop_index + 1:03d}.xlsx"
                )
                excel_report = LowVoltageExcelReport(
                    report_path,
                    configuration,
                )
            if dmm is not None:
                measurement_path = output_directory / (
                    f"low_voltage_dmm_measurements_loop_{loop_index + 1:03d}.csv"
                )
                measurement_file = measurement_path.open(
                    "w", newline="", encoding="utf-8"
                )
                measurement_writer = csv.DictWriter(
                    measurement_file,
                    fieldnames=(
                        "Loop",
                        "DUT Channel",
                        "DIAG Point",
                        "DMM Voltage (V)",
                        "Screenshot",
                    ),
                )
                measurement_writer.writeheader()
                measurement_file.flush()

            for channel in channels:
                self._checkpoint(worker)
                diag_channel = diag_channel_selector(channel)
                #dut.instr.write("*CLS")
                dut.senseVoltageSource("EXT", 1)
                dut.setMode("VOLTAGE", channel)
                dut.sourVoltageLevelImmediateAmplitude(0, channel)
                dut.outputState("ON", channel)
                enabled_channels.append(channel)
                dut.diag_POKE_LOW_VOLTAGE_CONTROL(
                    diag_channel,
                    points[0],
                )
                self._check_dut_error(dut, points[0])

                for index, point in enumerate(points):
                    self._checkpoint(worker)
                    if index:
                        dut.diag_POKE_LOW_VOLTAGE_CONTROL(
                            diag_channel,
                            point,
                        )
                        self._check_dut_error(dut, point)

                    self._emit(
                        worker,
                        f"Low-voltage capture: channel {channel}, "
                        f"point {point}",
                    )
                    self._wait(worker, delay)
                    dmm_voltage = self._measure_dmm(dmm, configuration)
                    if dmm_voltage is not None:
                        self._emit(
                            worker,
                            f"DMM voltage: {dmm_voltage:.9g} V",
                        )
                    scope.run()
                    self._wait(worker, scope_run_delay)
                    scope.stop()
                    scope_minimum, scope_maximum, scope_vpp = (
                        self._measure_scope_voltage(scope, configuration)
                    )
                    if scope_vpp is not None:
                        self._emit(
                            worker,
                            f"Scope VPP: {scope_vpp:.6g} mV",
                        )
                    image_data = strip_ieee_binary_block(
                        scope.read_binary_data()
                    )
                    image_path = output_directory / (
                        f"loop_{loop_index + 1:03d}_channel_{channel}_"
                        f"point_{point}.png"
                    )
                    image_path.write_bytes(image_data)
                    captured_images.append(image_path)
                    record = {
                        "Loop": loop_index + 1,
                        "DUT Channel": channel,
                        "DIAG Point": point,
                        "DMM Voltage (V)": dmm_voltage,
                        "Scope VMIN (V)": scope_minimum,
                        "Scope VMAX (V)": scope_maximum,
                        "Scope VPP (mV)": scope_vpp,
                        "Screenshot File": str(image_path),
                    }
                    if measurement_writer is not None:
                        measurement_writer.writerow(
                            {
                                "Loop": loop_index + 1,
                                "DUT Channel": channel,
                                "DIAG Point": point,
                                "DMM Voltage (V)": dmm_voltage,
                                "Screenshot": str(image_path),
                            }
                        )
                        measurement_file.flush()
                    if excel_report is not None:
                        excel_report.append(record, image_path)
                    completed += 1
                    if worker is not None:
                        worker.progress_value.emit(
                            int(completed * 100 / total)
                        )
                    self._emit(worker, f"Screenshot saved: {image_path}")

            return tuple(captured_images)
        finally:
            if excel_report is not None:
                try:
                    excel_report.save()
                    self._emit(worker, f"Excel report saved: {excel_report.path}")
                except Exception as exception:
                    self._emit(worker, f"Unable to save Excel report: {exception}")
            if measurement_file is not None:
                measurement_file.close()
            if scope is not None:
                try:
                    scope.stop()
                except Exception:
                    pass
            if dut is not None:
                for channel in enabled_channels:
                    try:
                        dut.outputState("OFF", channel)
                    except Exception:
                        pass
            for instrument in (dmm, scope, dut):
                session = getattr(instrument, "instr", None)
                if session is not None:
                    try:
                        session.close()
                    except Exception:
                        pass


class LowVoltageTestWorker(QThread):
    progress = pyqtSignal(str)
    progress_value = pyqtSignal(int)
    completed = pyqtSignal(object)
    stopped = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, configuration, runner=None):
        super().__init__()
        self.configuration = dict(configuration)
        self.runner = runner or LowVoltageScopeCaptureTest()
        self.run_context = None
        self._stop_requested = threading.Event()

    def stop(self):
        self._stop_requested.set()

    def checkpoint(self):
        if self._stop_requested.is_set():
            raise LowVoltageTestStopped(
                "Low-voltage test stopped by operator"
            )

    def interruptible_sleep(self, seconds):
        if self._stop_requested.wait(max(0.0, float(seconds))):
            self.checkpoint()

    def run(self):
        try:
            images = self.runner.execute(
                self.configuration,
                worker=self,
            )
            self.completed.emit(images)
        except LowVoltageTestStopped:
            self.stopped.emit()
        except Exception as exception:
            self.progress.emit(
                f"ERROR: {exception}\n{traceback.format_exc()}"
            )
            self.error.emit(str(exception))


class LowVoltageTestDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Hornbill Low Voltage Scope Capture")
        self.resize(760, 700)
        self.worker = None

        self.dut_address = QLineEdit()
        self.dut_address.setPlaceholderText("Hornbill DUT VISA address")
        self.scope_address = QLineEdit()
        self.scope_address.setPlaceholderText("Oscilloscope VISA address")
        self.dmm_enabled = QCheckBox("Enable DMM measurement at every point")
        self.dmm_enabled.setChecked(False)
        self.dmm_address = QLineEdit()
        self.dmm_address.setPlaceholderText("Optional DMM VISA address")
        self.dmm_model = QComboBox()
        self.dmm_model.addItem("344xxA / 34470A", "344xxA")
        self.dmm_model.addItem("3458A", "3458A")
        self.dmm_enabled.toggled.connect(self._update_dmm_controls)

        self.dut_channel = QSpinBox()
        self.dut_channel.setRange(1, 4)
        self.scope_channel = QSpinBox()
        self.scope_channel.setRange(1, 16)

        self.initial_point = QSpinBox()
        self.initial_point.setRange(0, 1_000_000)
        self.initial_point.setValue(DEFAULT_INITIAL_POINT)
        self.final_point = QSpinBox()
        self.final_point.setRange(0, 1_000_000)
        self.final_point.setValue(DEFAULT_FINAL_POINT)
        self.increment = QSpinBox()
        self.increment.setRange(1, 1_000_000)
        self.increment.setValue(LOW_VOLTAGE_INCREMENT)

        self.settling_delay = QDoubleSpinBox()
        self.settling_delay.setRange(0, 3600)
        self.settling_delay.setDecimals(3)
        self.settling_delay.setValue(1.0)
        self.settling_delay.setSuffix(" s")
        self.scope_run_delay = QDoubleSpinBox()
        self.scope_run_delay.setRange(0.1, 3600)
        self.scope_run_delay.setDecimals(3)
        self.scope_run_delay.setValue(3.0)
        self.scope_run_delay.setSuffix(" s")
        self.ink_saver = QComboBox()
        self.ink_saver.addItems(("ON", "OFF"))
        self.ink_saver.setToolTip(
            "ON uses the oscilloscope ink-saving hardcopy background; "
            "OFF keeps the normal display colors."
        )
        self.scope_vpp_enabled = QCheckBox(
            "Measure oscilloscope VPP using VMAX - VMIN"
        )
        self.scope_vpp_enabled.setChecked(True)
        self.excel_report_enabled = QCheckBox(
            "Create Excel report with embedded screenshots"
        )
        self.excel_report_enabled.setChecked(True)

        self.output_directory = QLineEdit()
        browse_button = QPushButton("Browse...")
        browse_button.clicked.connect(self._browse_output_directory)
        output_layout = QHBoxLayout()
        output_layout.addWidget(self.output_directory)
        output_layout.addWidget(browse_button)

        form = QFormLayout()
        form.addRow("Hornbill DUT Address:", self.dut_address)
        form.addRow("Oscilloscope Address:", self.scope_address)
        form.addRow(self.dmm_enabled)
        form.addRow("DMM Model:", self.dmm_model)
        form.addRow("DMM Address:", self.dmm_address)
        form.addRow("Hornbill Channel:", self.dut_channel)
        form.addRow("Oscilloscope Channel:", self.scope_channel)
        form.addRow("Initial DIAG Point:", self.initial_point)
        form.addRow("Final DIAG Point:", self.final_point)
        form.addRow("DIAG Increment:", self.increment)
        form.addRow("DIAG Settling Delay:", self.settling_delay)
        form.addRow("Scope Run Capture Delay:", self.scope_run_delay)
        form.addRow("Scope Ink Saver:", self.ink_saver)
        form.addRow(self.scope_vpp_enabled)
        form.addRow(self.excel_report_enabled)
        form.addRow("Output Folder:", output_layout)

        self.start_button = QPushButton("Start Capture")
        self.stop_button = QPushButton("Stop")
        self.stop_button.setEnabled(False)
        self.start_button.clicked.connect(self._start)
        self.stop_button.clicked.connect(self._stop)
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.stop_button)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.status = QLabel(
            "Ready. One oscilloscope PNG is captured at every selected "
            "DIAG increment."
        )
        self.status.setWordWrap(True)
        self.log = QTextEdit()
        self.log.setReadOnly(True)

        layout = QVBoxLayout(self)
        layout.addLayout(form)
        layout.addLayout(button_layout)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.status)
        layout.addWidget(self.log, stretch=1)
        self._update_dmm_controls(False)

    def _update_dmm_controls(self, enabled):
        self.dmm_model.setEnabled(bool(enabled))
        self.dmm_address.setEnabled(bool(enabled))

    def _browse_output_directory(self):
        directory = QFileDialog.getExistingDirectory(
            self,
            "Select Screenshot Output Folder",
        )
        if directory:
            self.output_directory.setText(directory)

    def _configuration(self):
        dut_address = self.dut_address.text().strip()
        scope_address = self.scope_address.text().strip()
        dmm_enabled = self.dmm_enabled.isChecked()
        dmm_address = self.dmm_address.text().strip()
        output_directory = self.output_directory.text().strip()
        if not dut_address or not scope_address:
            raise ValueError(
                "Hornbill DUT and oscilloscope addresses are required"
            )
        if dut_address == scope_address:
            raise ValueError(
                "Hornbill DUT and oscilloscope addresses must be different"
            )
        if dmm_enabled and not dmm_address:
            raise ValueError("DMM address is required when DMM measurement is enabled")
        if dmm_enabled and dmm_address in {dut_address, scope_address}:
            raise ValueError("DMM address must be different from DUT and oscilloscope")
        if not output_directory:
            raise ValueError("Select an output folder")
        low_voltage_points(
            self.initial_point.value(),
            self.final_point.value(),
            self.increment.value(),
        )
        return {
            "PSU": dut_address,
            "OSC": scope_address,
            "DMM_Enabled": dmm_enabled,
            "DMM": dmm_address if dmm_enabled else "",
            "DMM_Model": self.dmm_model.currentData(),
            "PSU_Channel": self.dut_channel.value(),
            "OSC_Channel": self.scope_channel.value(),
            "LowVoltageInitialPoint": self.initial_point.value(),
            "LowVoltageFinalPoint": self.final_point.value(),
            "LowVoltageIncrement": self.increment.value(),
            "updatedelay": self.settling_delay.value(),
            "ScopeRunCaptureDelay": self.scope_run_delay.value(),
            "InkSaver": self.ink_saver.currentText(),
            "ScopeVppEnabled": self.scope_vpp_enabled.isChecked(),
            "ExcelReportEnabled": self.excel_report_enabled.isChecked(),
            "savedir": output_directory,
        }

    def _start(self):
        try:
            configuration = self._configuration()
        except ValueError as exception:
            QMessageBox.warning(self, "Invalid Settings", str(exception))
            return

        self.log.clear()
        self.progress_bar.setValue(0)
        self.worker = LowVoltageTestWorker(configuration)
        self.worker.progress.connect(self.log.append)
        self.worker.progress_value.connect(self.progress_bar.setValue)
        self.worker.completed.connect(self._completed)
        self.worker.stopped.connect(self._stopped)
        self.worker.error.connect(self._error)
        self.worker.finished.connect(self._finished)
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.status.setText("Low-voltage scope capture is running...")
        self.worker.start()

    def _stop(self):
        if self.worker is not None:
            self.worker.stop()
            self.stop_button.setEnabled(False)
            self.status.setText("Stop requested; waiting for a safe point...")

    def _completed(self, images):
        self.progress_bar.setValue(100)
        measurement_note = (
            " DMM measurements were saved to CSV."
            if self.dmm_enabled.isChecked()
            else ""
        )
        report_note = (
            " An Excel report with embedded waveforms was generated."
            if self.excel_report_enabled.isChecked()
            else ""
        )
        self.status.setText(
            f"Completed. Captured {len(images)} oscilloscope image(s)."
            f"{measurement_note}{report_note}"
        )

    def _stopped(self):
        self.status.setText("Low-voltage scope capture stopped safely.")

    def _error(self, message):
        self.status.setText("Low-voltage scope capture failed.")
        QMessageBox.critical(self, "Low Voltage Test Error", message)

    def _finished(self):
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.worker = None

    def closeEvent(self, event):
        if self.worker is not None and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait(5000)
        super().closeEvent(event)
