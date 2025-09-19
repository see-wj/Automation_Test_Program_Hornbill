import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
import pandas as pd
import datetime
import os


class xlreportpower:
    """Generates an Excel report for voltage/current accuracy tests with conditional formatting and embedded images."""

    def __init__(self, save_directory, file_name=None):
        """Initialize formatting parameters and set file path for Excel report."""
        self.red_font = Font(size=14, bold=True, color="ffffff")
        self.red_fill = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")
        self.green_font = Font(size=14, bold=True, color="013220")
        self.green_fill = PatternFill(start_color="FFAAFF00", end_color="FFAAFF00", fill_type="solid")

        # Ensure the save directory exists
        if not os.path.exists(save_directory):
            os.makedirs(save_directory)

        # Generate a timestamped filename
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d--%H-%M-%S")
        if file_name:
            full_file_name = f"{file_name}_{timestamp}.xlsx"
        else:
            full_file_name = f"{timestamp}.xlsx"

        self.path = os.path.join(save_directory, full_file_name)

    def adjust_column_width(self, worksheet, width=20):
        """Set column width for columns A to R."""
        for col in range(1, 25):  # A to X
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    def align_cell(self, worksheet, row, col, style="center"):
        """Align a specific cell in the worksheet."""
        worksheet.cell(row=row, column=col).alignment = Alignment(horizontal=style)

    def run(self):
        """Generate the Excel report by importing CSV data, adding conditional formatting, and embedding an image."""
        # Define chart_path for image insertion
        current_dir = os.path.dirname(os.path.abspath(__file__))  # src/
        project_root = os.path.abspath(os.path.join(current_dir, ".."))
        csv_folder = os.path.join(project_root, "csv")
        chart_path = os.path.join(csv_folder, "images", "powerChart.png")

        try:
            with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
                # Read CSV files
                try:
                    df1 = pd.read_csv(os.path.join(csv_folder, "powererror.csv"))
                    df2 = pd.read_csv(os.path.join(csv_folder, "powerinstrumentData.csv"))
                    df4 = pd.read_csv(os.path.join(csv_folder, "powerconfig.csv"))
                except FileNotFoundError as e:
                    print(f"Error: {e}")
                    return

                # Write DataFrames to Excel
                df1.to_excel(writer, sheet_name="Data", index=False, startrow=7, startcol=3)
                df2.to_excel(writer, sheet_name="Data", index=False)
                df4.to_excel(writer, sheet_name="Data", index=False, startrow=6)

                # Access workbook and sheet
                wb = writer.book
                ws = wb["Data"]

                # Apply conditional formatting to df1's "Condition" column
                last_row = len(df1) + 8  # Assuming data starts at row 9
                cell_range1 = f"N9:N{last_row}"
                ws.conditional_formatting.add(
                    cell_range1,
                    FormulaRule(formula=['NOT(ISERROR(SEARCH("PASS",N9)))'], fill=self.green_fill, font=self.green_font),
                )
                ws.conditional_formatting.add(
                    cell_range1,
                    FormulaRule(formula=['NOT(ISERROR(SEARCH("FAIL",N9)))'], fill=self.red_fill, font=self.red_font),
                )
                
                cell_range2 = f"Q9:Q{last_row}"
                ws.conditional_formatting.add(
                    cell_range2,
                    FormulaRule(formula=['NOT(ISERROR(SEARCH("PASS",Q9)))'], fill=self.green_fill, font=self.green_font),
                )
                ws.conditional_formatting.add(
                    cell_range2,
                    FormulaRule(formula=['NOT(ISERROR(SEARCH("FAIL",Q9)))'], fill=self.red_fill, font=self.red_font),
                )


                # Adjust column widths
                self.adjust_column_width(ws, width=25)

                # Add timestamp to the sheet
                ws.cell(row=7, column=4, value="Time Generated:")
                ws.cell(row=7, column=5, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

                # Insert chart image, if it exists
                if os.path.exists(chart_path):
                    img = openpyxl.drawing.image.Image(chart_path)
                    img.anchor = "A34"  # Set image position
                    ws.add_image(img)
                else:
                    print(f"Warning: Image '{chart_path}' not found. Skipping image insertion.")

                # Save the workbook
                wb.save(self.path)

            print(f"Excel report saved: {self.path}")

        except Exception as e:
            import traceback
            traceback.print_exc()



