import os
import tempfile
import unittest
import zipfile
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.units import pixels_to_EMU

from reporting import data as data_module
from reporting import xlreport as xlreport_module
from SCPI_Library.simulation import reset_simulation
from execution.run_storage import create_run_storage


class ReportGenerationTests(unittest.TestCase):
    def setUp(self):
        reset_simulation()

    def _create_storage(self, base_directory, mode):
        storage = create_run_storage(base_directory, f"{mode}_SIMULATION")
        data_module.configure_run_storage(storage.raw, storage.charts)
        (storage.root / "SIMULATION_RUN.txt").write_text(
            "Simulated report integration test.\n", encoding="utf-8"
        )
        pd.DataFrame(
            {
                "Parameter": ["simulation_mode", "unit"],
                "Value": [True, mode],
            }
        ).to_csv(storage.raw / "config.csv", index=False)
        return storage

    def _assert_report_artifacts(
        self,
        storage,
        report_prefix,
        condition_ranges=4,
    ):
        expected_raw_files = {
            "config.csv",
            "data.csv",
            "error.csv",
            "instrumentData.csv",
        }
        self.assertTrue(expected_raw_files.issubset(
            {path.name for path in storage.raw.iterdir()}
        ))
        self.assertTrue((storage.charts / "Chart.png").is_file())
        self.assertTrue((storage.charts / "Chart2.png").is_file())
        self.assertTrue((storage.root / "SIMULATION_RUN.txt").is_file())

        reports = list(storage.reports.glob(f"{report_prefix}_*.xlsx"))
        self.assertEqual(len(reports), 1)
        workbook = load_workbook(reports[0])
        worksheet = workbook["Data"]
        self.assertEqual(len(worksheet._images), 2)
        self.assertIn("Charts", workbook.sheetnames)
        self.assertIn("ChartData", workbook.sheetnames)
        self.assertEqual(len(workbook["Charts"]._charts), 4)
        self.assertEqual(workbook["ChartData"].sheet_state, "hidden")
        charts = workbook["Charts"]._charts
        self.assertTrue(all(len(chart.series) >= 5 for chart in charts))
        for chart in charts:
            self.assertTrue(chart.title.tx.rich.p[0].pPr.defRPr.b)
            self.assertTrue(chart.x_axis.title.tx.rich.p[0].pPr.defRPr.b)
            self.assertTrue(chart.y_axis.title.tx.rich.p[0].pPr.defRPr.b)
            self.assertTrue(chart.x_axis.txPr.p[0].pPr.defRPr.b)
            self.assertTrue(chart.y_axis.txPr.p[0].pPr.defRPr.b)
            self.assertTrue(chart.legend.txPr.p[0].pPr.defRPr.b)
            self.assertGreaterEqual(
                chart.series[0].graphicalProperties.line.width,
                pixels_to_EMU(3.2),
            )
            self.assertGreaterEqual(
                chart.series[1].graphicalProperties.line.width,
                pixels_to_EMU(2.5),
            )
            self.assertGreaterEqual(
                max(
                    series.graphicalProperties.line.width or 0
                    for series in chart.series
                ),
                pixels_to_EMU(4.5),
            )
        self.assertEqual(len(worksheet.conditional_formatting), condition_ranges)
        values = {
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        self.assertIn("simulation_mode", values)

    def test_generates_simulated_voltage_report(self):
        with tempfile.TemporaryDirectory() as temporary_directory, patch.dict(
            os.environ, {"AUTOMATION_SIMULATION": "1"}, clear=False
        ):
            storage = self._create_storage(temporary_directory, "VOLTAGE")
            info = [[5.0, 1.0, 0], [6.0, 2.0, 0]]
            measured = [[5.0, 0.0], [7.0, 0.0]]
            readback = [[5.0, 1.0], [6.0, 2.0]]

            data_module.instrumentData(
                "USB0::SIM::PSU::INSTR",
                "USB0::SIM::DMM::INSTR",
                "USB0::SIM::ELOAD::INSTR",
            )
            graph = data_module.datatoGraph(info, measured, readback)
            graph.scatterCompareVoltage(
                0.001, 0.001, 0.001, 0.001, "VOLTAGE", 5.0
            )
            report = xlreport_module.xlreport(
                storage.reports, "SIMULATED_VOLTAGE"
            )
            report.run()

            data_columns = pd.read_csv(storage.raw / "data.csv").columns.tolist()
            self.assertEqual(
                data_columns[:4],
                [
                    "PSU Voltage Set",
                    "Load Current Set",
                    "PSU Readback Voltage",
                    "PSU Readback Current",
                ],
            )
            report_path = next(storage.reports.glob("SIMULATED_VOLTAGE_*.xlsx"))
            with zipfile.ZipFile(report_path) as workbook_archive:
                chart_xml = "".join(
                    workbook_archive.read(name).decode("utf-8")
                    for name in workbook_archive.namelist()
                    if name.startswith("xl/charts/chart")
                )
            self.assertIn("00B050", chart_xml)
            self.assertIn("FF0000", chart_xml)

            workbook = load_workbook(report_path)
            first_chart = workbook["Charts"]._charts[0]
            first_color = (
                first_chart.series[0]
                .graphicalProperties.line.solidFill.srgbClr
            )
            second_color = (
                first_chart.series[1]
                .graphicalProperties.line.solidFill.srgbClr
            )
            self.assertNotEqual(first_color, second_color)

            self._assert_report_artifacts(storage, "SIMULATED_VOLTAGE")

    def test_local_voltage_error_compares_vloc_with_vmon(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            storage = self._create_storage(temporary_directory, "VOLTAGE")
            info = [[5.0, 1.0, 0]]
            measured = [[5.0, 0.0]]
            readback = [[5.1, 1.0, 5.4]]

            data_module.datatoCSV_Accuracy(info, measured, readback)

            exported = pd.read_csv(storage.raw / "data.csv")
            self.assertAlmostEqual(
                exported.loc[0, "PSU Local Voltage Error (V)"],
                0.3,
            )

    def test_voltage_report_adds_local_voltage_error_excel_chart(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            storage = self._create_storage(temporary_directory, "VOLTAGE")
            info = [[5.0, 1.0, 0], [6.0, 1.0, 0]]
            measured = [[5.0, 0.0], [6.0, 0.0]]
            readback = [[5.1, 1.0, 5.4], [6.2, 1.0, 6.6]]

            data_module.instrumentData(
                "USB0::SIM::PSU::INSTR",
                "USB0::SIM::DMM::INSTR",
                "USB0::SIM::ELOAD::INSTR",
            )
            graph = data_module.datatoGraph(info, measured, readback)
            graph.scatterCompareVoltage(
                0.001, 0.001, 0.001, 0.001, "VOLTAGE", 5.0
            )
            report = xlreport_module.xlreport(
                storage.reports, "LOCAL_VOLTAGE_ERROR"
            )
            report.run()

            report_path = next(
                storage.reports.glob("LOCAL_VOLTAGE_ERROR_*.xlsx")
            )
            workbook = load_workbook(report_path)
            charts = workbook["Charts"]._charts
            self.assertEqual(len(charts), 5)
            self.assertEqual(len(workbook["Data"]._images), 2)
            self.assertGreaterEqual(len(charts[-1].series), 2)

            with zipfile.ZipFile(report_path) as workbook_archive:
                chart_xml = "".join(
                    workbook_archive.read(name).decode("utf-8")
                    for name in workbook_archive.namelist()
                    if name.startswith("xl/charts/chart")
                )
            self.assertIn("PSU Local Voltage Error (VLOC - VMON)", chart_xml)

    def test_voltage_current_change_chart_uses_load_current_as_x_axis(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            storage = self._create_storage(temporary_directory, "VOLTAGE")
            info = [[5.0, current, 0] for current in (1.0, 2.0, 3.0)]
            measured = [[5.0, 0.0] for _current in (1.0, 2.0, 3.0)]
            readback = [[5.0, current] for current in (1.0, 2.0, 3.0)]

            data_module.instrumentData(
                "USB0::SIM::PSU::INSTR",
                "USB0::SIM::DMM::INSTR",
                "USB0::SIM::ELOAD::INSTR",
            )
            graph = data_module.datatoGraph(info, measured, readback)
            graph.scatterCompareVoltage_Current_Change(
                0.001, 0.001, 0.001, 0.001, "VOLTAGE", 5.0
            )
            report = xlreport_module.xlreport(
                storage.reports, "SIMULATED_VOLTAGE_CURRENT_CHANGE"
            )
            report.run()

            chart_data = pd.read_csv(storage.raw / "error_percent.csv")
            self.assertEqual(
                chart_data["Chart Mode"].unique().tolist(),
                ["VOLTAGE_CURRENT_CHANGE"],
            )
            self.assertEqual(chart_data["Chart X"].tolist(), [1.0, 2.0, 3.0])
            self.assertEqual(
                chart_data["Chart Group"].unique().tolist(),
                ["Voltage = 5.0"],
            )
            self._assert_report_artifacts(
                storage,
                "SIMULATED_VOLTAGE_CURRENT_CHANGE",
                condition_ranges=2,
            )

    def test_generates_simulated_current_report(self):
        with tempfile.TemporaryDirectory() as temporary_directory, patch.dict(
            os.environ, {"AUTOMATION_SIMULATION": "1"}, clear=False
        ):
            storage = self._create_storage(temporary_directory, "CURRENT")
            info = [[5.0, 2.0, 0]]
            measured = [[0.0, 2.0]]
            readback = [[0.0, 2.0]]

            data_module.instrumentData(
                "USB0::SIM::PSU::INSTR",
                "USB0::SIM::DMM2::INSTR",
                "USB0::SIM::ELOAD::INSTR",
            )
            graph = data_module.datatoGraph2(info, measured, readback)
            graph.scatterCompareCurrent2(
                0.001, 0.001, 0.001, 0.001, "CURRENT", 2.0
            )
            report = xlreport_module.xlreport(
                storage.reports, "SIMULATED_CURRENT"
            )
            report.run()

            self._assert_report_artifacts(storage, "SIMULATED_CURRENT")

    def test_current_percentage_graph_sweeps_current_horizontally(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            storage = self._create_storage(temporary_directory, "CURRENT")
            info = [[1.0, current, 0] for current in (1.0, 2.0, 3.0)]
            measured = [[0.0, current] for current in (1.0, 2.0, 3.0)]
            readback = [[1.0, current] for current in (1.0, 2.0, 3.0)]
            graph = data_module.datatoGraph2(info, measured, readback)

            with patch(
                "matplotlib.axes.Axes.plot",
                autospec=True,
                return_value=[],
            ) as plot:
                graph.scatterCompareCurrent2(
                    0.001,
                    0.001,
                    0.001,
                    0.001,
                    "CURRENT",
                    10.0,
                )

            voltage_series = [
                call
                for call in plot.call_args_list
                if call.kwargs.get("label") == "Voltage = 1.0"
            ]

        self.assertEqual(len(voltage_series), 4)
        for call in voltage_series:
            self.assertEqual(list(call.args[1]), [1.0, 2.0, 3.0])


if __name__ == "__main__":
    unittest.main()
