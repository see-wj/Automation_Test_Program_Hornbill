import base64
import json
import tempfile
import unittest
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook

from instruments.scope_image_ocr import (
    ScopeImageOcrDialog,
    compile_scope_image_folder,
    decode_ocr_result_line,
    discover_scope_images,
    parse_display_value,
    parse_scope_filename,
)


class FakeOcrBackend:
    LABELS = {
        1: "DC RMS Cyc(4):",
        2: "Pk-Pk(4):",
        3: "Freq(3):",
        4: "Pk-Pk(3):",
        5: "DC RMS Cyc(3):",
        6: "DC RMS Cyc(2):",
        7: "Pk-Pk(2):",
        8: "Phase(1->3):",
    }
    VALUES = {
        1: "150mV",
        2: "1.04V",
        3: "500.1 kHz",
        4: "3.95V",
        5: "785mV",
        6: "2.41mV",
        7: "14.1mV",
        8: "108.84deg",
    }

    def __call__(self, items, stop_requested=None):
        results = {}
        for item in items:
            _image_index, slot, kind = item["id"].split(":")
            slot = int(slot)
            text = self.LABELS[slot] if kind == "label" else self.VALUES[slot]
            results[item["id"]] = {
                "id": item["id"],
                "text": text,
                "error": None,
            }
        return results


class ScopeImageOcrTests(unittest.TestCase):
    def test_decodes_ocr_text_containing_control_characters(self):
        expected = {
            "id": "0:1:value",
            "text": "150mV\nnoise\t\x0bmarker",
            "error": None,
        }
        encoded = base64.b64encode(
            json.dumps(expected).encode("utf-8")
        ).decode("ascii")

        self.assertEqual(expected, decode_ocr_result_line(encoded))

    def test_parses_scope_filename_metadata(self):
        first, second, timestamp = parse_scope_filename(
            "10.0_2.0_2026_07_24_00_27_09.png"
        )

        self.assertEqual("10.0", first)
        self.assertEqual("2.0", second)
        self.assertEqual("2026-07-24 00:27:09", timestamp)

    def test_normalizes_ocr_number_and_si_unit(self):
        display, numeric, unit, si_value, si_unit = parse_display_value(
            "500. I kHz"
        )

        self.assertEqual("500.1kHz", display)
        self.assertEqual(500.1, numeric)
        self.assertEqual("kHz", unit)
        self.assertEqual(500100.0, si_value)
        self.assertEqual("Hz", si_unit)

    def test_normalizes_ocr_space_as_decimal_separator(self):
        display, numeric, unit, si_value, si_unit = parse_display_value(
            "1 04V"
        )

        self.assertEqual("1.04V", display)
        self.assertEqual(1.04, numeric)
        self.assertEqual("V", unit)
        self.assertEqual(1.04, si_value)
        self.assertEqual("V", si_unit)

    def test_normalizes_common_ocr_digit_substitutions(self):
        display, numeric, unit, si_value, si_unit = parse_display_value(
            "I.IOV"
        )

        self.assertEqual("1.10V", display)
        self.assertEqual(1.1, numeric)
        self.assertEqual("V", unit)
        self.assertEqual(1.1, si_value)
        self.assertEqual("V", si_unit)

    def test_discovers_images_but_skips_generated_output(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            (root / "capture.png").write_bytes(b"image")
            excluded = root / "scope_ocr_analysis"
            excluded.mkdir()
            (excluded / "generated.png").write_bytes(b"image")

            self.assertEqual(
                (root / "capture.png",),
                discover_scope_images(root),
            )

    def test_compiles_measurements_into_excel(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            image_path = root / "10.0_2.0_2026_07_24_00_27_09.png"
            image = np.zeros((630, 800, 3), dtype=np.uint8)
            cv2.imwrite(str(image_path), image)
            output_path = root / "scope_measurements.xlsx"

            summary = compile_scope_image_folder(
                root,
                output_path,
                ocr_backend=FakeOcrBackend(),
            )

            self.assertEqual(1, summary.image_count)
            self.assertEqual(8, summary.measurement_count)
            self.assertEqual(0, summary.partial_count)
            workbook = load_workbook(output_path)
            summary_sheet = workbook["Summary"]
            measurements = workbook["Measurements"]
            self.assertEqual("Result", summary_sheet["A1"].value)
            self.assertEqual("Voltage", summary_sheet["B1"].value)
            self.assertEqual("Current", summary_sheet["C1"].value)
            self.assertEqual("Power (W)", summary_sheet["D1"].value)
            self.assertEqual("DC RMS Cycle", summary_sheet["E1"].value)
            self.assertEqual("Pk-Pk(4)", summary_sheet["F1"].value)
            self.assertEqual("Freq(3)", summary_sheet["G1"].value)
            self.assertEqual("Phase(1->3)", summary_sheet["L1"].value)
            self.assertEqual("Oscilloscope Image", summary_sheet["M1"].value)
            self.assertEqual(1, summary_sheet["A2"].value)
            self.assertEqual(10.0, summary_sheet["B2"].value)
            self.assertEqual(2.0, summary_sheet["C2"].value)
            self.assertEqual("=B2*C2", summary_sheet["D2"].value)
            self.assertEqual("150mV", summary_sheet["E2"].value)
            self.assertEqual("500.1kHz", summary_sheet["G2"].value)
            self.assertAlmostEqual(0.0141, summary_sheet["N2"].value)
            self.assertAlmostEqual(108.84, summary_sheet["O2"].value)
            self.assertTrue(summary_sheet.column_dimensions["N"].hidden)
            self.assertTrue(summary_sheet.column_dimensions["O"].hidden)
            self.assertEqual(1, len(summary_sheet._images))
            self.assertGreater(summary_sheet.row_dimensions[2].height, 30)
            self.assertEqual("DC RMS Cycle", measurements["G2"].value)
            self.assertEqual("150mV", measurements["I2"].value)
            self.assertAlmostEqual(0.15, measurements["L2"].value)
            self.assertEqual("Frequency", measurements["G4"].value)
            self.assertAlmostEqual(500100.0, measurements["L4"].value)
            self.assertEqual("Phase", measurements["G9"].value)
            charts_sheet = workbook["Charts"]
            self.assertEqual(2, len(charts_sheet._charts))
            self.assertEqual("Power Relationship Charts", charts_sheet["A1"].value)

    def test_summary_sorts_low_voltage_then_low_current(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            image = np.zeros((630, 800, 3), dtype=np.uint8)
            filenames = (
                "10.0_5.0_2026_07_24_00_00_03.png",
                "5.0_20.0_2026_07_24_00_00_02.png",
                "5.0_1.0_2026_07_24_00_00_01.png",
            )
            for filename in filenames:
                cv2.imwrite(str(root / filename), image)

            output_path = root / "scope_measurements.xlsx"
            compile_scope_image_folder(
                root,
                output_path,
                ocr_backend=FakeOcrBackend(),
            )
            summary_sheet = load_workbook(output_path)["Summary"]

            self.assertEqual(
                [(5.0, 1.0), (5.0, 20.0), (10.0, 5.0)],
                [
                    (summary_sheet.cell(row, 2).value, summary_sheet.cell(row, 3).value)
                    for row in range(2, 5)
                ],
            )

    def test_separates_repeated_condition_captures_by_result(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            image = np.zeros((630, 800, 3), dtype=np.uint8)
            first = root / "10.0_2.0_2026_08_05_19_51_35.png"
            second = root / "10.0_2.0_2026_08_05_19_51_41.png"
            cv2.imwrite(str(first), image)
            cv2.imwrite(str(second), image)

            output_path = root / "scope_measurements.xlsx"
            compile_scope_image_folder(
                root,
                output_path,
                ocr_backend=FakeOcrBackend(),
            )
            workbook = load_workbook(output_path)
            summary_sheet = workbook["Summary"]

            self.assertEqual([1, 2], [summary_sheet["A2"].value, summary_sheet["A3"].value])
            self.assertIn("Result 1", workbook.sheetnames)
            self.assertIn("Result 2", workbook.sheetnames)
            self.assertEqual(2, workbook["Result 1"].max_row)
            self.assertEqual(2, workbook["Result 2"].max_row)
            self.assertTrue(workbook["Result 1"]["M2"].value.endswith(first.name))
            self.assertTrue(workbook["Result 2"]["M2"].value.endswith(second.name))
            self.assertEqual(2, len(workbook["Charts"]._charts[0].series))

    def test_main_test_selection_registers_scope_ocr_dialog(self):
        import GUI

        registry = GUI.MainWindow._create_dialog_registry(object())
        registration = next(
            item
            for item in registry.registrations
            if item.owner_attribute == "scope_image_ocr_dialog"
        )

        self.assertEqual("Oscilloscope Image OCR to Excel", registration.title)
        self.assertIs(ScopeImageOcrDialog, registration.factory)


if __name__ == "__main__":
    unittest.main()
